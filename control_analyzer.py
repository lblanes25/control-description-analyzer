#!/usr/bin/env python3
"""
Control Description Analyzer

This module analyzes control descriptions for completeness and quality.
It identifies the essential elements of controls (WHO, WHAT, WHEN, WHY, ESCALATION)
and detects when multiple controls are combined into a single description.

The analyzer leverages NLP techniques to evaluate control descriptions and
provides detailed analysis results with improvement suggestions.
"""

import argparse
import os
import sys
import re
import pickle
import gc
import logging
import time
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple, Union

import pandas as pd
import numpy as np
import spacy
from spacy.matcher import PhraseMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Config and support modules
from config_manager import ConfigManager

# Import enhanced detection modules
from enhanced_who import enhanced_who_detection_v2
from enhanced_what import enhance_what_detection, mark_possible_standalone_controls
from enhanced_when import enhance_when_detection
from enhanced_why import enhance_why_detection
from enhanced_escalation import enhance_escalation_detection
from enhanced_multi_control import detect_multi_control

# Set up logging
logger = logging.getLogger("control_analyzer")

logging.basicConfig(level=logging.DEBUG)

class ControlElement:
    """
    Class representing a single control element with keywords and scoring logic.

    Attributes:
        name: The name of the element (WHO, WHAT, etc.)
        weight: The weight of this element in scoring (0-100)
        keywords: List of keywords related to this element
        score: Raw score from analysis
        matched_keywords: Keywords matched during analysis
        enhanced_results: Results from enhanced detection modules
        normalized_score: Score normalized to 0-100 scale
        context_relevance: How relevant keywords are in context
        matcher: spaCy PhraseMatcher for this element
        phrases: spaCy Doc objects for keywords
    """

    def __init__(self, name: str = None, weight: int = 0, keywords: List[str] = None):
        """
        Initialize a control element with name, weight and keywords.

        Args:
            name: The name of the element (WHO, WHAT, etc.)
            weight: The weight of this element in scoring (0-100)
            keywords: List of keywords related to this element
        """
        self.name = name
        self.weight = weight
        self.keywords = keywords or []
        self.score = 0
        self.matched_keywords = []
        self.enhanced_results = {}
        self.normalized_score = 0
        self.context_relevance = 0
        self.matcher = None
        self.phrases = []

    def setup_matchers(self, nlp):
        """
        Set up phrase matchers for this element's keywords.

        Args:
            nlp: spaCy NLP model
        """
        self.matcher = PhraseMatcher(nlp.vocab, attr="LOWER")

        # Add single-word and multi-word phrases
        self.phrases = [nlp(keyword) for keyword in self.keywords]
        if self.phrases:
            self.matcher.add(f"{self.name}_patterns", self.phrases)

    def analyze(self, text: str, nlp, enhanced_mode: bool = True, **context) -> float:
        """
        Analyze text for presence of this element's keywords.

        Args:
            text: The text to analyze
            nlp: spaCy NLP model
            enhanced_mode: Whether to use enhanced detection modules
            **context: Additional context like control_type, frequency, etc.

        Returns:
            Weighted score for this element
        """
        if not text or text.strip() == '':
            return 0

        # Use enhanced detection if available
        if enhanced_mode:
            if self.name == "WHO":
                control_type = context.get("control_type")
                frequency = context.get("frequency")
                self.enhanced_results = enhanced_who_detection_v2(
                    text, nlp, control_type, frequency, self.keywords)
                self.score = self.enhanced_results.get("confidence", 0) if self.enhanced_results else 0
                self.normalized_score = self.score * 100
                self.matched_keywords = [
                    self.enhanced_results.get("primary", {}).get("text", "")
                ] if self.enhanced_results and self.enhanced_results.get("primary") else []

            elif self.name == "WHAT":
                self.enhanced_results = enhance_what_detection(text, nlp, self.keywords)
                self.score = self.enhanced_results.get("score", 0) if self.enhanced_results else 0
                self.normalized_score = self.score * 100
                if self.enhanced_results and self.enhanced_results.get("primary_action"):
                    self.matched_keywords = [self.enhanced_results["primary_action"]["full_phrase"]]
                else:
                    self.matched_keywords = [a["full_phrase"] for a in self.enhanced_results.get("actions", [])[
                                                                       :3]] if self.enhanced_results else []

            elif self.name == "WHEN":
                self.enhanced_results = enhance_when_detection(
                    text,
                    nlp,
                    context.get("control_type"),
                    self.keywords,
                    context.get("frequency")
                )
                self.score = self.enhanced_results.get("score", 0) if self.enhanced_results else 0
                self.normalized_score = self.score * 100
                self.matched_keywords = (self.enhanced_results.get("extracted_keywords", [])
                                         if self.enhanced_results else [])

            elif self.name == "WHY":
                risk_description = context.get("risk_description")

                # Store original text for comparison
                original_text = text.strip()

                self.enhanced_results = enhance_why_detection(
                    text, nlp, risk_description, self.keywords)
                self.score = self.enhanced_results.get("score", 0) if self.enhanced_results else 0
                self.normalized_score = self.score * 100

                # Initialize to empty list by default
                self.matched_keywords = []

                if self.enhanced_results:
                    # Check for explicit top match
                    top = self.enhanced_results.get("top_match", {})
                    if isinstance(top, dict) and top:
                        # Check for implied purpose first
                        if "implied_purpose" in top:
                            implied = top["implied_purpose"]
                            # Only use if it's not the same as the original text
                            if implied and implied.lower() != original_text.lower():
                                self.matched_keywords = [implied]
                        # Then check for explicit text
                        elif "text" in top:
                            top_text = top["text"]
                            # Only use if it's not the same as the original text
                            if top_text and top_text.lower() != original_text.lower():
                                self.matched_keywords = [top_text]

                    # If still empty, try extracted keywords
                    if not self.matched_keywords:
                        extracted = self.enhanced_results.get("extracted_keywords", [])
                        # Filter out any that match the original text
                        filtered = [k for k in extracted if k.lower() != original_text.lower()]
                        if filtered:
                            self.matched_keywords = filtered

            elif self.name == "ESCALATION":
                self.enhanced_results = enhance_escalation_detection(text, nlp, self.keywords)
                self.score = self.enhanced_results.get("score", 0) if self.enhanced_results else 0
                self.normalized_score = self.score * 100
                self.matched_keywords = [p["text"] for p in
                                         self.enhanced_results.get("phrases", [])] if self.enhanced_results else []

            else:
                # Fall back to base implementation for other elements
                return self._base_analyze(text, nlp)

            # Return weighted score
            return self.score * self.weight
        else:
            # Use base implementation
            return self._base_analyze(text, nlp)

    def _base_analyze(self, text: str, nlp) -> float:
        """
        Original analysis method as fallback.

        Args:
            text: Text to analyze
            nlp: spaCy NLP model

        Returns:
            Weighted score for this element
        """
        # Create spaCy doc
        doc = nlp(text.lower())

        # Find matches using PhraseMatcher
        matches = self.matcher(doc)

        # Extract matched keywords
        matched_spans = []
        for match_id, start, end in matches:
            span = doc[start:end]
            matched_spans.append((span, span.text))

        # Get unique matched keywords
        self.matched_keywords = list(set(match[1] for match in matched_spans))

        # Calculate context relevance
        if matched_spans:
            # Assess the syntactic importance of matched words
            relevance_scores = []
            for span, _ in matched_spans:
                # Check if span is part of a subject or verb phrase
                for token in span:
                    if token.dep_ in ("nsubj", "dobj", "ROOT", "VERB"):
                        relevance_scores.append(1.0)  # High relevance
                    else:
                        relevance_scores.append(0.7)  # Medium relevance

            self.context_relevance = sum(relevance_scores) / len(relevance_scores) if relevance_scores else 0.5
        else:
            self.context_relevance = 0.0

        # Calculate raw score (0 to 1, based on matches and context)
        self.score = min(1.0, len(self.matched_keywords) * 0.25) if self.matched_keywords else 0

        # Apply context relevance
        self.score = self.score * max(0.7, self.context_relevance)

        # Return weighted score
        self.normalized_score = self.score * 100
        return self.score * self.weight

    def get_enhancement_feedback(self) -> Optional[Union[str, List[str]]]:
        """
        Get feedback from enhanced detection if available.

        Returns:
            Feedback string or list, or None if not available
        """
        if not self.enhanced_results:
            return None

        if self.name == "WHO":
            return self.enhanced_results.get("message")
        elif self.name == "WHAT":
            return self.enhanced_results.get("suggestions", [])
        elif self.name == "WHEN":
            return self.enhanced_results.get("improvement_suggestions", [])
        elif self.name == "WHY":
            return self.enhanced_results.get("feedback")
        elif self.name == "ESCALATION":
            return self.enhanced_results.get("suggestions", [])
        else:
            return None


def enhanced_semantic_similarity(text: str, element_keywords: List[str], nlp) -> Tuple[List[Tuple], float]:
    """
    Enhanced semantic similarity detection that better captures contextual meaning.

    Args:
        text: The text to analyze
        element_keywords: List of keywords to check similarity against
        nlp: spaCy NLP model

    Returns:
        Tuple of (similar terms found, average similarity score)
    """
    if not text or text.strip() == '':
        return [], 0.0

    # Process the text
    doc = nlp(text.lower())

    # Create a list to store similar terms found
    similar_terms = []
    similarity_scores = []

    # Analyze at chunk level for better context
    for chunk in doc.noun_chunks:
        # Skip very short chunks
        if len(chunk) < 2:
            continue

        # Get vector for this chunk
        if not chunk.has_vector:
            continue

        # Check similarity against all keywords and keyword phrases
        for keyword in element_keywords:
            # Skip if exact match (this is already handled by exact matcher)
            if keyword.lower() in chunk.text.lower():
                continue

            # Get keyword vector - handle phrases properly
            keyword_doc = nlp(keyword.lower())

            # Skip if no vector
            if not keyword_doc.has_vector:
                continue

            # Calculate similarity
            similarity = chunk.similarity(keyword_doc)

            # Only consider moderately high similarities to avoid false positives
            if similarity > 0.65:
                similar_terms.append((chunk.text, keyword, similarity))
                similarity_scores.append(similarity)

    # Also analyze at sentence level for more context
    for sent in doc.sents:
        if not sent.has_vector:
            continue

        # Compare each sentence with concatenated keywords to detect thematic similarity
        keyword_text = " ".join(element_keywords[:10])  # Use first 10 keywords to avoid too long text
        keyword_doc = nlp(keyword_text)

        if keyword_doc.has_vector:
            sent_similarity = sent.similarity(keyword_doc)

            if sent_similarity > 0.7:  # Higher threshold for sentence-level to avoid false positives
                similar_terms.append((sent.text, "element theme", sent_similarity))
                similarity_scores.append(sent_similarity * 0.8)  # Slight discount for full sentences

    # Sort by similarity score
    similar_terms.sort(key=lambda x: x[2], reverse=True)

    # Calculate average similarity score
    avg_similarity = sum(similarity_scores) / len(similarity_scores) if similarity_scores else 0

    return similar_terms[:5], avg_similarity  # Limit to top 5 to avoid noise


def boost_term_by_context(term: str, doc) -> float:
    """
    Boost similarity scores based on the term's role in the sentence.

    Args:
        term: The term to analyze
        doc: spaCy Doc containing the term

    Returns:
        Boost factor (multiplicative)
    """
    boost = 1.0

    # Find the term in the document
    term_tokens = []
    for token in doc:
        if token.text.lower() in term.lower():
            term_tokens.append(token)

    if not term_tokens:
        return boost

    # Apply boosts based on dependency relations
    for token in term_tokens:
        # Subject terms are more important
        if token.dep_ in ("nsubj", "nsubjpass"):
            boost *= 1.3

        # Direct objects of main verbs are important
        elif token.dep_ == "dobj" and token.head.dep_ == "ROOT":
            boost *= 1.2

        # Terms that are direct children of root verbs
        elif token.head.dep_ == "ROOT":
            boost *= 1.15

        # Terms in prepositional phrases are less important
        elif token.dep_ == "pobj":
            boost *= 0.9

    return min(1.5, boost)  # Cap the boost


def disambiguate_control_term(term: str, context: str, nlp) -> bool:
    """
    Disambiguate whether a term like 'check' or 'record' is being used in a control-relevant sense.

    Args:
        term: The term to disambiguate
        context: Surrounding context of the term
        nlp: spaCy NLP model

    Returns:
        Boolean indicating if the term is used in a control-relevant sense
    """
    doc = nlp(context)
    term = term.lower()

    # Only support known ambiguous terms
    if term not in {"check", "record"}:
        return True

    # Define disallowed objects that suggest non-control use
    non_control_objects = {
        "check": {"payment", "mark", "number"},
        "record": {"time", "number", "volume"}
    }

    # Define allowed control-related objects (optional, for boosting)
    control_objects = {
        "check": {"accuracy", "completeness", "compliance", "validity"},
        "record": {"results", "evidence", "findings", "control"}
    }

    for token in doc:
        if token.lemma_ == term and token.pos_ in {"VERB", "NOUN"}:
            for child in token.children:
                if child.dep_ in {"dobj", "pobj", "attr"}:
                    obj = child.lemma_.lower()

                    if obj in non_control_objects.get(term, set()):
                        return False  # Likely non-control usage
                    if obj in control_objects.get(term, set()):
                        return True  # Strong control-related usage

    # Fall back to string-based disambiguation if needed
    context_lower = context.lower()
    for phrase in control_objects.get(term, []):
        if phrase in context_lower:
            return True
    for phrase in non_control_objects.get(term, []):
        if phrase in context_lower:
            return False

    return True  # Default to assuming control-relevance


class EnhancedControlAnalyzer:
    """
    Enhanced analyzer with specialized detection modules for each element.

    This class coordinates the analysis of control descriptions using
    a combination of NLP techniques and specialized detection modules.
    """

    def __init__(self, config_file: Optional[str] = None):
        """
        Initialize the analyzer with configuration settings.

        Args:
            config_file: Optional path to configuration YAML file
        """
        # Configure logging
        self._setup_logging()

        # Load config manager
        self.config_manager = ConfigManager(config_file)
        self.config = self.config_manager.config if self.config_manager else {}

        # Get column mappings from config YAML file
        self.column_mappings = self._initialize_column_mappings()

        # Initialize spaCy with model specified in config
        self.nlp = self._initialize_spacy_model()

        # Initialize elements with their weights from config
        self.elements = self._initialize_elements()

        # Set up matchers for each element
        for element in self.elements.values():
            element.setup_matchers(self.nlp)

        # Vague terms that should be avoided - loaded from config
        self.vague_terms = self.config.get('vague_terms', [])
        self.vague_matcher = PhraseMatcher(self.nlp.vocab, attr="LOWER")
        vague_phrases = [self.nlp(term) for term in self.vague_terms]
        if vague_phrases:
            self.vague_matcher.add("vague_patterns", vague_phrases)

        # Configure enhanced detection - can be overridden via config
        self.use_enhanced_detection = self.config.get('use_enhanced_detection', True)

        # Get penalty configuration from config
        self.vague_term_penalty = self.config.get('vague_term_penalty', 2)
        self.max_vague_penalty = self.config.get('max_vague_penalty', 10)

        # Get multi-control penalty settings from config
        self.multi_control_penalty = self.config.get('penalties', {}).get('multi_control', {})
        self.points_per_control = self.multi_control_penalty.get('points_per_control', 5)
        self.max_multi_control_penalty = self.multi_control_penalty.get('max_penalty', 10)

        # Get domain clusters from config
        self.domain_clusters = self.config.get('domain_clusters', {})

        # Get thresholds from config
        self.excellent_threshold = self.config.get('category_thresholds', {}).get('excellent', 75)
        self.good_threshold = self.config.get('category_thresholds', {}).get('good', 50)

        # Get audit leader column from config
        self.audit_leader_column = self.config_manager.get_audit_leader_column()

        # Get audit entity column from config
        self.audit_entity_column = self.config_manager.get_audit_entity_column()

        logger.info("Analyzer initialized successfully")

    def _setup_logging(self):
        """Configure the logging system for the analyzer."""
        # Set up root logger
        root_logger = logging.getLogger()
        if not root_logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            handler.setFormatter(formatter)
            root_logger.addHandler(handler)
            root_logger.setLevel(logging.INFO)

        # Set module logger level
        logger.setLevel(logging.INFO)

    def _initialize_column_mappings(self) -> Dict[str, str]:
        """
        Initialize column mappings from config.

        Returns:
            Dictionary mapping internal column keys to actual Excel column names
        """
        column_mappings = {}

        # Get mappings from config if available
        if self.config_manager:
            yaml_columns = self.config_manager.get_column_defaults()
            if yaml_columns:
                logger.info(f"Loaded column mappings from YAML: {yaml_columns}")
                column_mappings = yaml_columns
            else:
                logger.info("No column mappings found in YAML file, using defaults")

        # Log the mappings
        logger.debug("Using column mappings:")
        for key, value in column_mappings.items():
            logger.debug(f"  {key}: {value}")

        return column_mappings

    def _initialize_spacy_model(self):
        """
        Initialize spaCy model based on configuration.

        Returns:
            Loaded spaCy model
        """
        spacy_config = self.config.get('spacy', {})
        preferred_model = spacy_config.get('preferred_model', 'en_core_web_md')
        fallback_model = spacy_config.get('fallback_model', 'en_core_web_sm')

        try:
            # Try loading the preferred model first
            return spacy.load(preferred_model)
        except OSError:
            try:
                # Fall back to smaller model if preferred isn't available
                nlp = spacy.load(fallback_model)
                logger.warning(
                    f"Using {fallback_model} model. For better results, install {preferred_model} "
                    f"with: python -m spacy download {preferred_model}"
                )
                return nlp
            except OSError:
                logger.info(f"Downloading spaCy model {fallback_model}...")
                spacy.cli.download(fallback_model)
                return spacy.load(fallback_model)

    def _get_column_name(self, column_key: str, override_value: Optional[str] = None) -> str:
        """
        Get column name from config or use override value or default.

        Args:
            column_key: Internal column key to look up
            override_value: Optional override value

        Returns:
            Actual column name to use
        """
        # Default column mappings for fallback
        default_column_mappings = {
            "id": "Control ID",
            "description": "Control Description",
            "frequency": "Control Frequency",
            "type": "Control Type",
            "risk": "Key Risk Description",
            "audit_leader": "Audit Leader from AE",
            "audit_entity": "Audit Entity"
        }

        if override_value:
            return override_value

        # Try from config mappings first
        if column_key in self.column_mappings:
            return self.column_mappings[column_key]

        # Fall back to default mapping
        return default_column_mappings.get(column_key)

    def _initialize_elements(self) -> Dict[str, ControlElement]:
        """
        Initialize elements with weights and keywords from configuration.

        Returns:
            Dictionary of ControlElement objects
        """
        elements = {}
        config_elements = self.config.get('elements', {})

        # Default element configuration if not specified in YAML
        default_elements = {
            "WHO": {"weight": 32, "keywords": []},
            "WHEN": {"weight": 22, "keywords": []},
            "WHAT": {"weight": 32, "keywords": []},
            "WHY": {"weight": 11, "keywords": []},
            "ESCALATION": {"weight": 3, "keywords": []}
        }

        # Merge defaults with configuration
        for name, defaults in default_elements.items():
            element_config = config_elements.get(name, {})
            weight = element_config.get('weight', defaults['weight'])
            keywords = element_config.get('keywords', defaults['keywords'])

            elements[name] = ControlElement(name, weight, keywords)

        return elements

    def analyze_control(self, control_id: str, description: str,
                        frequency: Optional[str] = None,
                        control_type: Optional[str] = None,
                        risk_description: Optional[str] = None) -> Dict[str, Any]:
        """
        Analyze a single control description and return detailed results.
        Enhanced with specialized detection modules for each element.

        Args:
            control_id: Unique identifier for the control
            description: The control description text
            frequency: Optional declared frequency for validation
            control_type: Optional control type for context
            risk_description: Optional risk description for WHY alignment

        Returns:
            Dictionary with comprehensive analysis results
        """
        if not description or pd.isna(description):
            return {
                "control_id": control_id,
                "description": "",
                "total_score": 0,
                "category": "Needs Improvement",
                "missing_elements": list(self.elements.keys()),
                "vague_terms_found": [],
                "weighted_scores": {element: 0 for element in self.elements},
                "matched_keywords": {element: [] for element in self.elements},
                "enhancement_feedback": {element: None for element in self.elements},
                "validation_results": {
                    "frequency_valid": False,
                    "frequency_message": "No description provided",
                    "control_type_valid": False,
                    "control_type_message": "No description provided"
                },
                "multi_control_indicators": {"detected": False, "count": 0, "candidates": []}
            }

        # Reset all elements for new analysis
        for element in self.elements.values():
            element.score = 0
            element.matched_keywords = []
            element.enhanced_results = {}

        # Create context for element analysis
        context = {
            "control_type": control_type,
            "frequency": frequency,
            "risk_description": risk_description
        }

        # Analyze for each element with specialized detection
        normalized_scores = {}
        weighted_scores = {}
        matched_keywords = {}
        enhancement_feedback = {}

        for name, element in self.elements.items():
            element.analyze(
                description,
                self.nlp,
                self.use_enhanced_detection,
                **context
            )
            normalized_scores[name] = element.normalized_score  # For reporting
            weight = element.weight
            weighted_scores[name] = (element.normalized_score * weight / 100)  # Contribution to total
            matched_keywords[name] = element.matched_keywords
            enhancement_feedback[name] = element.get_enhancement_feedback()

        # Now that all elements have been analyzed, check for multi-control description
        # This ensures we have the element analysis results to feed into multi-control detection
        multi_control_indicators = self.detect_multi_control(description)

        # Check for vague terms with PhraseMatcher
        doc = self.nlp(description)
        vague_matches = self.vague_matcher(doc)
        vague_terms_found = []

        for match_id, start, end in vague_matches:
            span = doc[start:end]
            vague_terms_found.append(span.text)

        # Remove duplicates
        vague_terms_found = list(set(vague_terms_found))

        # Calculate vague terms penalty based on config
        vague_penalty = min(len(vague_terms_found) * self.vague_term_penalty, self.max_vague_penalty)

        # Calculate total score
        total_score = sum(weighted_scores.values()) - vague_penalty
        total_score = max(0, total_score)  # Ensure score is not negative

        # Apply multi-control penalty if detected, using config values
        if multi_control_indicators["detected"]:
            multi_control_penalty = min(
                self.max_multi_control_penalty,
                multi_control_indicators["count"] * self.points_per_control
            )
            total_score = max(0, total_score - multi_control_penalty)

        # Determine category based on config thresholds
        if total_score >= self.excellent_threshold:
            category = "Excellent"
        elif total_score >= self.good_threshold:
            category = "Good"
        else:
            category = "Needs Improvement"

        # Define threshold by element type
        element_thresholds = {
            "WHO": 5.0,
            "WHEN": 5.0,
            "WHAT": 5.0,
            "WHY": 2.0,  # Lower threshold for WHY since it typically scores lower
            "ESCALATION": 1.0  # Lowest threshold as these are often bonus elements
        }

        # Identify missing elements using element-specific thresholds
        missing_elements = [name for name, score in weighted_scores.items()
                            if score < element_thresholds.get(name, 5.0)]

        # Add validation results
        validation_results = {}

        # Use enhanced WHEN validation results if available
        when_element = self.elements["WHEN"]
        if self.use_enhanced_detection and "validation" in when_element.enhanced_results:
            validation_results["frequency_valid"] = when_element.enhanced_results["validation"]["is_valid"]
            validation_results["frequency_message"] = when_element.enhanced_results["validation"]["message"]
        else:
            # Fallback to basic validation
            if frequency:
                frequency_valid, frequency_message = self.validate_frequency(description, frequency)
                validation_results["frequency_valid"] = frequency_valid
                validation_results["frequency_message"] = frequency_message
            else:
                validation_results["frequency_valid"] = False
                validation_results["frequency_message"] = "No frequency value provided"

        # Control type validation
        if control_type:
            control_type_valid, control_type_message = self.validate_control_type(description, control_type)
            validation_results["control_type_valid"] = control_type_valid
            validation_results["control_type_message"] = control_type_message
        else:
            validation_results["control_type_valid"] = False
            validation_results["control_type_message"] = "No control type value provided"

        return {
            "control_id": control_id,
            "description": description,
            "total_score": total_score,
            "category": category,
            "missing_elements": missing_elements,
            "vague_terms_found": vague_terms_found,
            "weighted_scores": weighted_scores,
            "normalized_scores": normalized_scores,
            "matched_keywords": matched_keywords,
            "enhancement_feedback": enhancement_feedback,
            "validation_results": validation_results,
            "multi_control_indicators": multi_control_indicators
        }

    def detect_multi_control(self, description: str) -> Dict[str, Any]:
        """
        Detect if a description contains multiple controls using enhanced detection
        that considers WHO, WHAT, WHEN and ESCALATION elements and their relationships.

        Args:
            description: The control description text

        Returns:
            Dictionary with detection results indicating if multiple controls are present
        """
        if not description or description.strip() == '':
            return {
                "detected": False,
                "count": 0,
                "candidates": []
            }

        # If enhanced detection is disabled, fall back to the simpler approach
        if not self.use_enhanced_detection:
            # Use the simpler what-based detection as fallback
            potential_controls = mark_possible_standalone_controls(description, self.nlp)

            return {
                "detected": len(potential_controls) > 1,
                "count": len(potential_controls),
                "candidates": potential_controls,
                "confidence": "low"
            }

        # For enhanced detection, we need results from individual element analyses
        # to properly detect control relationships

        # Get WHO data
        who_element = self.elements["WHO"]
        who_data = who_element.enhanced_results if hasattr(who_element, 'enhanced_results') else {}

        # Get WHAT data
        what_element = self.elements["WHAT"]
        what_data = what_element.enhanced_results if hasattr(what_element, 'enhanced_results') else {}

        # Get WHEN data
        when_element = self.elements["WHEN"]
        when_data = when_element.enhanced_results if hasattr(when_element, 'enhanced_results') else {}

        # Get ESCALATION data
        escalation_element = self.elements["ESCALATION"]
        escalation_data = escalation_element.enhanced_results if hasattr(escalation_element, 'enhanced_results') else {}

        # Get the configuration for multi-control detection
        multi_control_config = self.config.get("multi_control", {})

        # Use the enhanced multi-control detection
        multi_control_result = detect_multi_control(
            description,
            who_data,
            what_data,
            when_data,
            escalation_data,
            multi_control_config
        )

        # Return result with the expected structure
        result = {
            "detected": multi_control_result.get("detected", False),
            "count": multi_control_result.get("count", 0),
            "candidates": multi_control_result.get("candidates", []),
            "confidence": multi_control_result.get("confidence", "low"),
            "timing_groups": multi_control_result.get("timing_groups", []),
            "has_multi_frequency": multi_control_result.get("has_multi_frequency", False),
            "has_sequence_markers": multi_control_result.get("has_sequence_markers", False)
        }

        return result

    def validate_frequency(self, description: str, frequency_value: str) -> Tuple[bool, str]:
        """
        Validate that the frequency mentioned in the description matches the frequency field value.

        Args:
            description: Control description text
            frequency_value: Expected frequency value

        Returns:
            Tuple of (is_valid, validation_message)
        """
        if not description or not frequency_value:
            return False, "Missing description or frequency value"

        # Get frequency terms from config
        frequency_terms = self.config.get('frequency_terms', {})

        # Default frequencies if not in config
        default_frequency_terms = {
            "daily": ["daily", "each day", "every day"],
            "weekly": ["weekly", "each week", "every week"],
            "monthly": ["monthly", "each month", "every month"],
            "quarterly": ["quarterly", "each quarter", "every quarter"],
            "annually": ["annually", "yearly", "each year", "every year"]
        }

        # Merge with defaults if config doesn't have these frequencies
        for freq, terms in default_frequency_terms.items():
            if freq not in frequency_terms:
                frequency_terms[freq] = terms

        # Normalize frequency value to lowercase
        frequency_value = frequency_value.lower().strip()

        # Get standard terms for this frequency if available
        expected_terms = frequency_terms.get(frequency_value, [frequency_value])

        # Check if any expected terms are in the description
        description_lower = description.lower()
        matches = []

        for term in expected_terms:
            pattern = r'\b' + re.escape(term) + r'\b'
            if re.search(pattern, description_lower):
                matches.append(term)

        # If matches found, frequency is consistent
        if matches:
            return True, f"Frequency in description ({', '.join(matches)}) matches expected value ({frequency_value})"
        else:
            return False, f"Frequency in description does not match expected value ({frequency_value})"

    def validate_control_type(self, description: str, control_type: str) -> Tuple[bool, str]:
        """
        Validate that the control type mentioned in the description matches the control type field value.

        Args:
            description: Control description text
            control_type: Expected control type value

        Returns:
            Tuple of (is_valid, validation_message)
        """
        if not description or not control_type:
            return False, "Missing description or control type value"

        # Get control type keywords from config
        control_type_keywords = self.config.get('control_type_keywords', {})

        # Default control type keywords if not in config
        default_control_type_keywords = {
            "preventive": ["prevent", "preventive", "preventative", "avoid", "before", "prior to", "stops"],
            "detective": ["detect", "detective", "identify", "discover", "find", "after", "following", "review",
                          "monitor", "reconcile"],
            "corrective": ["correct", "corrective", "remediate", "fix", "resolve", "address", "adjust"],
            "automated": ["system", "automated", "automatic", "software", "program", "script",
                          "job", "batch", "workflow", "algorithm", "systematic"]
        }

        # Merge with defaults if config doesn't have these types
        for type_name, keywords in default_control_type_keywords.items():
            if type_name not in control_type_keywords:
                control_type_keywords[type_name] = keywords

        # Normalize control type
        control_type = control_type.lower().strip()

        # Get relevant keywords for this control type
        expected_keywords = control_type_keywords.get(control_type, [])

        if not expected_keywords:
            return False, f"Unknown control type: {control_type}"

        # Check if any expected keywords are in the description
        description_lower = description.lower()
        matches = []

        for keyword in expected_keywords:
            pattern = r'\b' + re.escape(keyword) + r'\b'
            if re.search(pattern, description_lower):
                matches.append(keyword)

        # If matches found, control type is consistent
        if matches:
            return True, f"Control type in description ({', '.join(matches)}) is consistent with {control_type}"
        else:
            return False, f"Control type in description does not indicate {control_type}"

    def _check_audit_entity_column(self, df: pd.DataFrame,
                                   audit_entity_column: Optional[str]) -> Optional[str]:
        """
        Check if audit entity column exists and try to auto-detect if not.

        Args:
            df: DataFrame with control data
            audit_entity_column: Specified audit entity column name

        Returns:
            Validated or auto-detected audit entity column name
        """
        if audit_entity_column and audit_entity_column not in df.columns:
            logger.warning(f"Audit Entity column '{audit_entity_column}' not found in file.")

            # Try to automatically detect Audit Entity column
            potential_columns = ["Audit Entity", "audit entity", "Audit_Entity", "audit_entity",
                                 "AuditEntity", "auditEntity", "Entity", "AE",
                                 "Business Unit", "Department"]

            for col in potential_columns:
                if col in df.columns:
                    audit_entity_column = col
                    logger.info(f"Automatically detected Audit Entity column: '{col}'")
                    break
            else:
                audit_entity_column = None

        return audit_entity_column

    def analyze_file(self, file_path: str, id_column: Optional[str] = None,
                     desc_column: Optional[str] = None, freq_column: Optional[str] = None,
                     type_column: Optional[str] = None, risk_column: Optional[str] = None,
                     audit_leader_column: Optional[str] = None,
                     audit_entity_column: Optional[str] = None,
                     output_file: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Analyze controls from an Excel file and generate a detailed report.

        Args:
            file_path: Path to Excel file containing controls
            id_column: Column containing control IDs (overrides config)
            desc_column: Column containing control descriptions (overrides config)
            freq_column: Column containing frequency values (overrides config)
            type_column: Column containing control type values (overrides config)
            risk_column: Column containing risk descriptions (overrides config)
            audit_leader_column: Column containing audit leader info (overrides config)
            audit_entity_column: Column containing audit entity info (overrides config)
            output_file: Optional path for output Excel report

        Returns:
            List of control analysis results
        """
        # Get column names from config (with overrides from parameters)
        id_column = self._get_column_name("id", id_column)
        desc_column = self._get_column_name("description", desc_column)
        freq_column = self._get_column_name("frequency", freq_column)
        type_column = self._get_column_name("type", type_column)
        risk_column = self._get_column_name("risk", risk_column)
        audit_leader_column = self._get_column_name("audit_leader", audit_leader_column)
        audit_entity_column = self._get_column_name("audit_entity", audit_entity_column)

        logger.info(f"Reading file: {file_path}")
        logger.info(f"Using columns: ID={id_column}, Description={desc_column}, "
                    f"Frequency={freq_column}, Type={type_column}, Risk={risk_column}, "
                    f"Audit Leader={audit_leader_column}, Audit Entity={audit_entity_column}")

        try:
            # Read the Excel file
            df = pd.read_excel(file_path, engine='openpyxl')

            # Ensure required columns exist
            if id_column not in df.columns:
                logger.error(f"ID column '{id_column}' not found in file")
                raise ValueError(f"ID column '{id_column}' not found in file")

            if desc_column not in df.columns:
                logger.error(f"Description column '{desc_column}' not found in file")
                raise ValueError(f"Description column '{desc_column}' not found in file")

            # Check optional columns
            if freq_column and freq_column not in df.columns:
                logger.warning(
                    f"Frequency column '{freq_column}' not found in file. "
                    "Frequency validation will be skipped."
                )
                freq_column = None

            if type_column and type_column not in df.columns:
                logger.warning(
                    f"Control type column '{type_column}' not found in file. "
                    "Control type validation will be skipped."
                )
                type_column = None

            if risk_column and risk_column not in df.columns:
                logger.warning(
                    f"Risk description column '{risk_column}' not found in file. "
                    "Risk alignment will be skipped."
                )
                risk_column = None

            # Check and try to auto-detect audit leader column
            audit_leader_column = self._check_audit_leader_column(df, audit_leader_column)

            # Check and try to auto-detect audit entity column
            audit_entity_column = self._check_audit_entity_column(df, audit_entity_column)

            # Analyze each control
            results = []
            total_controls = len(df)

            logger.info(f"Analyzing {total_controls} controls...")

            for i, (idx, row) in enumerate(df.iterrows()):
                # Show progress
                if i % 100 == 0 or i == total_controls - 1:
                    progress = (i + 1) / total_controls * 100
                    logger.info(f"Progress: {progress:.1f}% ({i + 1}/{total_controls})")

                # Process this control
                result = self._process_single_control(row, id_column, desc_column,
                                                      freq_column, type_column,
                                                      risk_column, audit_leader_column,
                                                      audit_entity_column)
                results.append(result)

            # Create output file if specified
            if output_file:
                self._generate_enhanced_report(
                    results,
                    output_file,
                    freq_column is not None,
                    type_column is not None,
                    risk_column is not None
                )
                logger.info(f"Analysis complete. Results saved to {output_file}")

            return results

        except Exception as e:
            logger.error(f"Error analyzing file: {e}", exc_info=True)
            raise

    def _check_audit_leader_column(self, df: pd.DataFrame,
                                   audit_leader_column: Optional[str]) -> Optional[str]:
        """
        Check if audit leader column exists and try to auto-detect if not.

        Args:
            df: DataFrame with control data
            audit_leader_column: Specified audit leader column name

        Returns:
            Validated or auto-detected audit leader column name
        """
        if audit_leader_column and audit_leader_column not in df.columns:
            logger.warning(f"Audit Leader column '{audit_leader_column}' not found in file.")

            # Try to automatically detect Audit Leader column
            potential_columns = ["Audit Leader", "audit leader", "Audit_Leader", "audit_leader",
                                 "AuditLeader", "auditLeader", "Auditor", "Lead Auditor"]

            for col in potential_columns:
                if col in df.columns:
                    audit_leader_column = col
                    logger.info(f"Automatically detected Audit Leader column: '{col}'")
                    break
            else:
                audit_leader_column = None

        return audit_leader_column

    def _process_single_control(self, row: pd.Series, id_column: str, desc_column: str,
                                freq_column: Optional[str] = None,
                                type_column: Optional[str] = None,
                                risk_column: Optional[str] = None,
                                audit_leader_column: Optional[str] = None,
                                audit_entity_column: Optional[str] = None) -> Dict[str, Any]:
        """
        Process a single control from a DataFrame row.

        Args:
            row: DataFrame row with control data
            id_column: Column containing control ID
            desc_column: Column containing control description
            freq_column: Column containing frequency value
            type_column: Column containing control type
            risk_column: Column containing risk description
            audit_leader_column: Column containing audit leader
            audit_entity_column: Column containing audit entity

        Returns:
            Control analysis result
        """
        control_id = row[id_column]
        description = row[desc_column]

        # Extract optional fields with null checks
        frequency = None
        if freq_column and freq_column in row:
            freq_value = row[freq_column]
            if pd.notna(freq_value):  # Check if not NaN
                frequency = str(freq_value) if not isinstance(freq_value, str) else freq_value

        control_type = None
        if type_column and type_column in row:
            type_value = row[type_column]
            if pd.notna(type_value):  # Check if not NaN
                control_type = str(type_value) if not isinstance(type_value, str) else type_value

        risk_description = None
        if risk_column and risk_column in row:
            risk_value = row[risk_column]
            if pd.notna(risk_value):  # Check if not NaN
                risk_description = str(risk_value) if not isinstance(risk_value, str) else risk_value

        # Analyze control
        result = self.analyze_control(control_id, description, frequency, control_type, risk_description)

        # Add Audit Leader if available
        if audit_leader_column and audit_leader_column in row:
            audit_leader = row[audit_leader_column]
            if not pd.isna(audit_leader):
                result["Audit Leader"] = audit_leader

        # Add Audit Entity if available
        if audit_entity_column and audit_entity_column in row:
            audit_entity = row[audit_entity_column]
            if not pd.isna(audit_entity):
                result["Audit Entity"] = audit_entity

        return result

    def analyze_file_with_batches(self, file_path: str, id_column: Optional[str] = None,
                                  desc_column: Optional[str] = None,
                                  freq_column: Optional[str] = None,
                                  type_column: Optional[str] = None,
                                  risk_column: Optional[str] = None,
                                  audit_leader_column: Optional[str] = None,
                                  audit_entity_column: Optional[str] = None,
                                  output_file: Optional[str] = None,
                                  batch_size: int = 500,
                                  temp_dir: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Analyze controls from an Excel file in batches and generate a detailed report.

        Args:
            file_path: Path to Excel file containing controls
            id_column: Column containing control IDs (overrides config)
            desc_column: Column containing control descriptions (overrides config)
            freq_column: Column containing frequency values (overrides config)
            type_column: Column containing control type values (overrides config)
            risk_column: Column containing risk descriptions (overrides config)
            audit_leader_column: Column containing audit leader info (overrides config)
            output_file: Optional path for output Excel report
            batch_size: Number of controls to process in each batch
            temp_dir: Directory to store temporary batch results

        Returns:
            List of control analysis results
        """
        # Get column names from config (with overrides from parameters)
        resolved_columns = self._resolve_column_names(
            id_column, desc_column, freq_column, type_column, risk_column,
            audit_leader_column, audit_entity_column
        )
        id_column = resolved_columns["id_column"]
        desc_column = resolved_columns["desc_column"]
        freq_column = resolved_columns["freq_column"]
        type_column = resolved_columns["type_column"]
        risk_column = resolved_columns["risk_column"]
        audit_leader_column = resolved_columns["audit_leader_column"]
        audit_entity_column = resolved_columns["audit_entity_column"]

        logger.info(f"Reading file: {file_path}")
        logger.info(f"Using columns: ID={id_column}, Description={desc_column}, "
                    f"Frequency={freq_column}, Type={type_column}, Risk={risk_column}, "
                    f"Audit Leader={audit_leader_column}, Audit Entity={audit_entity_column}")
        logger.info(f"Using batch size: {batch_size}")

        # Create temp directory if needed
        temp_dir = self._prepare_temp_directory(temp_dir)

        try:
            # Read the Excel file
            df = pd.read_excel(file_path, engine='openpyxl')

            # Validate columns
            self._validate_dataframe_columns(df, id_column, desc_column)

            # Check optional columns
            freq_column = self._check_optional_column(df, freq_column, "Frequency")
            type_column = self._check_optional_column(df, type_column, "Control type")
            risk_column = self._check_optional_column(df, risk_column, "Risk description")

            # Check and try to auto-detect audit leader column
            audit_leader_column = self._check_audit_leader_column(df, audit_leader_column)

            # Initialize results and tracking variables
            all_results = []
            total_controls = len(df)
            start_time = time.time()

            logger.info(f"Analyzing {total_controls} controls in batches of {batch_size}...")

            # Process in batches
            batch_results = self._process_in_batches(
                df, total_controls, batch_size, id_column, desc_column,
                freq_column, type_column, risk_column, audit_leader_column,
                temp_dir, all_results
            )

            # All batches processed - create final report
            if output_file:
                self._generate_final_report(
                    batch_results, output_file, freq_column, type_column, risk_column, temp_dir
                )

            # Calculate total time
            self._log_processing_summary(start_time, total_controls)

            return batch_results

        except Exception as e:
            logger.error(f"Error analyzing file: {e}", exc_info=True)
            self._handle_analysis_error(temp_dir, e)
            raise

    def _resolve_column_names(self, id_column, desc_column, freq_column,
                              type_column, risk_column, audit_leader_column,
                              audit_entity_column):
        """
        Resolve all column names from config or parameters.

        Args:
            Various column name parameters

        Returns:
            Dictionary with resolved column names
        """
        return {
            "id_column": self._get_column_name("id", id_column),
            "desc_column": self._get_column_name("description", desc_column),
            "freq_column": self._get_column_name("frequency", freq_column),
            "type_column": self._get_column_name("type", type_column),
            "risk_column": self._get_column_name("risk", risk_column),
            "audit_leader_column": self._get_column_name("audit_leader", audit_leader_column),
            "audit_entity_column": self._get_column_name("audit_entity", audit_entity_column)
        }

    def _prepare_temp_directory(self, temp_dir):
        """
        Create and return the temporary directory path.

        Args:
            temp_dir: Specified temp directory path or None

        Returns:
            Directory path to use
        """
        if temp_dir:
            os.makedirs(temp_dir, exist_ok=True)
        else:
            temp_dir = "temp_batch_results"
            os.makedirs(temp_dir, exist_ok=True)
        return temp_dir

    def _validate_dataframe_columns(self, df, id_column, desc_column):
        """
        Validate that required columns exist in the DataFrame.

        Args:
            df: DataFrame to check
            id_column: ID column name
            desc_column: Description column name

        Raises:
            ValueError: If required columns are missing
        """
        if id_column not in df.columns:
            logger.error(f"ID column '{id_column}' not found in file")
            raise ValueError(f"ID column '{id_column}' not found in file")

        if desc_column not in df.columns:
            logger.error(f"Description column '{desc_column}' not found in file")
            raise ValueError(f"Description column '{desc_column}' not found in file")

    def _check_optional_column(self, df, column_name, column_type):
        """
        Check if optional column exists and log warning if not.

        Args:
            df: DataFrame to check
            column_name: Column name to check
            column_type: Type of column for logging

        Returns:
            Column name if valid, None otherwise
        """
        if column_name and column_name not in df.columns:
            logger.warning(
                f"{column_type} column '{column_name}' not found in file. "
                f"{column_type} validation will be skipped."
            )
            return None
        return column_name

    def _process_in_batches(self, df, total_controls, batch_size, id_column, desc_column,
                            freq_column, type_column, risk_column, audit_leader_column,
                            temp_dir, all_results):
        """
        Process the DataFrame in batches.

        Args:
            Various parameters for batch processing

        Returns:
            Combined results from all batches
        """
        for batch_start in range(0, total_controls, batch_size):
            batch_end = min(batch_start + batch_size, total_controls)
            batch_num = batch_start // batch_size + 1
            total_batches = (total_controls + batch_size - 1) // batch_size

            batch_start_time = time.time()
            logger.info(f"\nProcessing batch {batch_num}/{total_batches}: "
                        f"controls {batch_start + 1}-{batch_end}")

            # Get the slice of DataFrame for this batch
            batch_df = df.iloc[batch_start:batch_end].copy()

            # Process this batch
            batch_results = self._process_batch(
                batch_df, batch_num, id_column, desc_column,
                freq_column, type_column, risk_column, audit_leader_column
            )

            # Save batch results to temporary file
            self._save_batch_results(batch_results, batch_num, temp_dir)

            # Add to overall results
            all_results.extend(batch_results)

            # Report batch statistics and save checkpoint
            self._report_batch_completion(
                batch_num, total_batches, batch_df, batch_start_time,
                batch_end, total_controls, all_results, temp_dir
            )

            # Clean up memory
            self._cleanup_memory()

        return all_results

    def _process_batch(self, batch_df, batch_num, id_column, desc_column,
                       freq_column, type_column, risk_column, audit_leader_column):
        """
        Process a single batch of controls.

        Args:
            batch_df: DataFrame slice for this batch
            Other parameters for control analysis

        Returns:
            List of analysis results for this batch
        """
        batch_results = []

        # Process each control in the batch
        for i, (idx, row) in enumerate(batch_df.iterrows()):
            if i % 25 == 0:
                # Show more granular progress within batch
                progress = (i + 1) / len(batch_df) * 100
                logger.info(f"  Batch progress: {progress:.1f}% ({i + 1}/{len(batch_df)})")

            try:
                # Process this control
                result = self._process_single_control(
                    row, id_column, desc_column, freq_column,
                    type_column, risk_column, audit_leader_column
                )
                batch_results.append(result)
            except Exception as e:
                # Log error but continue processing
                control_id = row[id_column]
                description = row[desc_column]
                logger.error(f"Error analyzing control {control_id}: {str(e)}")

                # Add minimal error entry to maintain indexing
                batch_results.append({
                    "control_id": control_id,
                    "description": description,
                    "total_score": 0,
                    "category": "Error",
                    "missing_elements": [],
                    "error_message": str(e)
                })

        return batch_results

    def _save_batch_results(self, batch_results, batch_num, temp_dir):
        """
        Save batch results to a temporary file.

        Args:
            batch_results: Results to save
            batch_num: Batch number
            temp_dir: Directory to save in
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        batch_filename = os.path.join(temp_dir, f"batch_{batch_num:04d}_{timestamp}.pkl")

        try:
            with open(batch_filename, 'wb') as f:
                pickle.dump(batch_results, f, protocol=4)  # Use protocol 4 for compatibility
            logger.info(f"  Batch results saved to {batch_filename}")
        except Exception as e:
            logger.warning(f"  Could not save batch file: {e}")

    def _report_batch_completion(self, batch_num, total_batches, batch_df,
                                 batch_start_time, batch_end, total_controls,
                                 all_results, temp_dir):
        """
        Report batch completion statistics and save checkpoint.

        Args:
            Various parameters for reporting batch completion
        """
        # Calculate batch statistics
        batch_time = time.time() - batch_start_time
        controls_per_second = len(batch_df) / batch_time if batch_time > 0 else 0

        # Estimate remaining time
        completed = batch_end
        remaining = total_controls - completed
        estimated_time = remaining / controls_per_second if controls_per_second > 0 else 0

        # Report batch completion
        logger.info(f"  Batch {batch_num}/{total_batches} completed in {batch_time:.1f} seconds")
        logger.info(f"  Processing speed: {controls_per_second:.2f} controls/second")
        logger.info(f"  Progress: {completed}/{total_controls} controls "
                    f"({completed / total_controls * 100:.1f}%)")

        hours, remainder = divmod(estimated_time, 3600)
        minutes, seconds = divmod(remainder, 60)
        logger.info(f"  Estimated time remaining: {int(hours)}h {int(minutes)}m {int(seconds)}s")

        # Save checkpoint
        self._save_checkpoint(all_results, temp_dir)

    def _save_checkpoint(self, all_results, temp_dir):
        """
        Save a checkpoint of all results processed so far.

        Args:
            all_results: Results to save
            temp_dir: Directory to save in
        """
        checkpoint_file = os.path.join(temp_dir, "all_results_checkpoint.pkl")
        try:
            with open(checkpoint_file, 'wb') as f:
                pickle.dump(all_results, f, protocol=4)
            logger.info(f"  Checkpoint saved to {checkpoint_file}")
        except Exception as e:
            logger.warning(f"  Could not save checkpoint: {e}")

    def _cleanup_memory(self):
        """Clean up memory between batches."""
        # Clear memory
        gc.collect()

        # Optional: Clear the NLP model's cache to free memory
        try:
            # Get all pipelines and clear their caches
            for pipe_name in self.nlp.pipe_names:
                pipe = self.nlp.get_pipe(pipe_name)
                if hasattr(pipe, "vocab") and hasattr(pipe.vocab, "strings"):
                    pipe.vocab.strings.clean_up()
        except Exception:
            # If cleanup fails, just continue
            pass

    def _generate_final_report(self, results, output_file, freq_column,
                               type_column, risk_column, temp_dir):
        """
        Generate the final report with error handling.

        Args:
            results: Analysis results
            output_file: Output file path
            freq_column: Frequency column name
            type_column: Control type column name
            risk_column: Risk column name
            temp_dir: Temp directory for emergency saves
        """
        try:
            self._generate_enhanced_report(
                results,
                output_file,
                freq_column is not None,
                type_column is not None,
                risk_column is not None
            )
            logger.info(f"\nAnalysis complete. Results saved to {output_file}")
        except Exception as e:
            logger.error(f"Error generating final report: {e}")

            # Try to save results in pickle format
            emergency_output = os.path.splitext(output_file)[0] + "_emergency.pkl"

            try:
                with open(emergency_output, 'wb') as f:
                    pickle.dump(results, f, protocol=4)
                logger.info(f"Emergency results saved to {emergency_output}")
            except Exception as save_error:
                logger.error(f"Error saving emergency results: {save_error}")

    def _log_processing_summary(self, start_time, total_controls):
        """
        Log a summary of the entire processing job.

        Args:
            start_time: Job start time
            total_controls: Total number of controls processed
        """
        total_time = time.time() - start_time
        hours, remainder = divmod(total_time, 3600)
        minutes, seconds = divmod(remainder, 60)

        logger.info("\n----- Processing Summary -----")
        logger.info(f"Total processing time: {int(hours)}h {int(minutes)}m {int(seconds)}s")
        logger.info(f"Average processing speed: {total_controls / total_time:.2f} controls/second")

    def _handle_analysis_error(self, temp_dir, error):
        """
        Handle errors during analysis by saving any results we have.

        Args:
            temp_dir: Temp directory for emergency saves
            error: The exception that occurred
        """
        # Try to save any results we have as an emergency backup
        emergency_file = os.path.join(
            temp_dir, f"emergency_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pkl"
        )

        if 'all_results' in locals() and locals()['all_results']:
            try:
                with open(emergency_file, 'wb') as f:
                    pickle.dump(locals()['all_results'], f, protocol=4)
                logger.info(f"Emergency results saved to {emergency_file}")
            except Exception as e:
                logger.error(f"Failed to save emergency results: {e}")

    def _generate_enhanced_report(self, results: List[Dict], output_file: str,
                                  include_frequency: bool = False,
                                  include_control_type: bool = False,
                                  include_risk_alignment: bool = False) -> bool:
        """
        Generate a detailed Excel report with the enhanced analysis results - improved version
        that prevents Excel corruption issues and handles errors gracefully.

        Args:
            results: List of control analysis results
            output_file: Path to save the Excel report
            include_frequency: Whether to include frequency validation
            include_control_type: Whether to include control type validation
            include_risk_alignment: Whether to include risk alignment analysis

        Returns:
            Boolean indicating success
        """
        try:
            # ========== PREPARE DATA FRAMES ==========
            report_data = self._prepare_report_data(
                results, include_frequency, include_control_type, include_risk_alignment
            )

            # ========== CALCULATE SUMMARY STATISTICS ==========
            summary_stats = self._calculate_summary_statistics(results)

            # ========== CREATE WORKBOOK AND SHEETS ==========
            wb = self._create_report_workbook(
                report_data, summary_stats, include_frequency,
                include_control_type, include_risk_alignment
            )

            # ========== FORMAT WORKBOOK ==========
            self._format_workbook_sheets(wb)

            # ========== SAVE WORKBOOK ==========
            return self._save_workbook(wb, output_file)

        except Exception as e:
            logger.error(f"Critical error generating report: {str(e)}", exc_info=True)
            return False

    def _prepare_report_data(self, results: List[Dict], include_frequency: bool,
                             include_control_type: bool, include_risk_alignment: bool) -> Dict:
        """
        Prepare DataFrames for the Excel report.

        Args:
            results: List of control analysis results
            include_frequency: Whether to include frequency validation
            include_control_type: Whether to include control type validation
            include_risk_alignment: Whether to include risk alignment analysis

        Returns:
            Dictionary containing prepared DataFrames
        """
        # Analysis Results DataFrame
        basic_results = []
        for r in results:
            # Get basic data with defaults for safety
            result_dict = {
                "Control ID": self._safe_get(r, "control_id", ""),
                "Description": self._safe_text(self._safe_get(r, "description", ""), 2000),
                "Total Score": self._safe_get(r, "total_score", 0),
                "Category": self._safe_get(r, "category", "Unknown"),
                "Missing Elements": self._safe_text(", ".join(self._safe_get(r, "missing_elements", []))
                                                    if self._safe_get(r, "missing_elements") else "None"),
                "Vague Terms": self._safe_text(", ".join(self._safe_get(r, "vague_terms_found", []))
                                               if self._safe_get(r, "vague_terms_found") else "None"),
                "WHO Score": self._safe_get(r, ["normalized_scores", "WHO"], 0),
                "WHEN Score": self._safe_get(r, ["normalized_scores", "WHEN"], 0),
                "WHAT Score": self._safe_get(r, ["normalized_scores", "WHAT"], 0),
                "WHY Score": self._safe_get(r, ["normalized_scores", "WHY"], 0),
                "ESCALATION Score": self._safe_get(r, ["normalized_scores", "ESCALATION"], 0),
            }

            # Add enhanced multi-control indicators
            multi_detected = self._safe_get(r, ["multi_control_indicators", "detected"], False)
            multi_count = self._safe_get(r, ["multi_control_indicators", "count"], 0)
            result_dict["Multiple Controls"] = f"Yes ({multi_count})" if multi_detected else "No"

            # Add additional multi-control information from enhanced detection
            if 'confidence' in r.get("multi_control_indicators", {}):
                result_dict["Multi-Control Confidence"] = r["multi_control_indicators"]["confidence"]

            if 'timing_groups' in r.get("multi_control_indicators", {}):
                groups = r["multi_control_indicators"]["timing_groups"]
                if isinstance(groups, list):
                    result_dict["Timing Groups"] = ", ".join(str(g) for g in groups)

            if 'has_multi_frequency' in r.get("multi_control_indicators", {}):
                result_dict["Multiple Frequencies"] = "Yes" if r["multi_control_indicators"][
                    "has_multi_frequency"] else "No"

            if 'has_sequence_markers' in r.get("multi_control_indicators", {}):
                result_dict["Sequence Markers"] = "Yes" if r["multi_control_indicators"][
                    "has_sequence_markers"] else "No"

            # Add validation results if applicable
            if include_frequency:
                freq_valid = self._safe_get(r, ["validation_results", "frequency_valid"], False)
                freq_message = self._safe_get(r, ["validation_results", "frequency_message"], "")
                result_dict["Frequency Valid"] = "Yes" if freq_valid else "No"
                result_dict["Frequency Message"] = self._safe_text(freq_message)

            if include_control_type:
                type_valid = self._safe_get(r, ["validation_results", "control_type_valid"], False)
                type_message = self._safe_get(r, ["validation_results", "control_type_message"], "")
                result_dict["Control Type Valid"] = "Yes" if type_valid else "No"
                result_dict["Control Type Message"] = self._safe_text(type_message)

            # Add risk alignment feedback if available
            if include_risk_alignment:
                why_feedback = self._safe_get(r, ["enhancement_feedback", "WHY"], None)
                if why_feedback:
                    result_dict["Risk Alignment Feedback"] = self._safe_text(why_feedback)

            basic_results.append(result_dict)

        df_results = pd.DataFrame(basic_results)

        # Keywords DataFrame
        keyword_results = []
        for r in results:
            kw_dict = {
                "Control ID": self._safe_get(r, "control_id", ""),
                "WHO Keywords": self._safe_text(
                    self._join_keywords(self._safe_get(r, ["matched_keywords", "WHO"], []))),
                "WHEN Keywords": self._safe_text(
                    self._join_keywords(self._safe_get(r, ["matched_keywords", "WHEN"], []))),
                "WHAT Keywords": self._safe_text(
                    self._join_keywords(self._safe_get(r, ["matched_keywords", "WHAT"], []))),
                "WHY Keywords": self._safe_text(
                    self._join_keywords(self._safe_get(r, ["matched_keywords", "WHY"], []))),
                "ESCALATION Keywords": self._safe_text(
                    self._join_keywords(self._safe_get(r, ["matched_keywords", "ESCALATION"], [])))
            }
            keyword_results.append(kw_dict)

        df_keywords = pd.DataFrame(keyword_results)

        # Enhancement Feedback DataFrame - the problematic one
        feedback_results = []
        for r in results:
            fb_dict = {"Control ID": self._safe_get(r, "control_id", "")}

            # Process each element's feedback carefully
            for element in ["WHO", "WHEN", "WHAT", "WHY", "ESCALATION"]:
                feedback = self._safe_get(r, ["enhancement_feedback", element], None)

                # Format based on type, with extra safety
                if isinstance(feedback, list):
                    # Join list items with semicolons, max length 500
                    fb_dict[f"{element} Feedback"] = self._safe_text(
                        "; ".join(str(item) for item in feedback), 500)
                elif isinstance(feedback, str):
                    # Limit string length to 500
                    fb_dict[f"{element} Feedback"] = self._safe_text(feedback, 500)
                else:
                    fb_dict[f"{element} Feedback"] = "None"

            feedback_results.append(fb_dict)

        df_feedback = pd.DataFrame(feedback_results)

        # Return all prepared DataFrames
        return {
            "df_results": df_results,
            "df_keywords": df_keywords,
            "df_feedback": df_feedback
        }

    def _calculate_summary_statistics(self, results: List[Dict]) -> Dict:
        """
        Calculate summary statistics from analysis results.

        Args:
            results: List of control analysis results

        Returns:
            Dictionary of summary statistics
        """
        # Basic counts
        total_controls = len(results)
        excellent_count = sum(1 for r in results if self._safe_get(r, "category") == "Excellent")
        good_count = sum(1 for r in results if self._safe_get(r, "category") == "Good")
        needs_improvement_count = sum(1 for r in results if self._safe_get(r, "category") == "Needs Improvement")

        # Average score
        scores = [self._safe_get(r, "total_score", 0) for r in results]
        avg_score = sum(scores) / max(len(scores), 1)

        # Missing elements counts
        missing_elements_counts = {element: 0 for element in self.elements}
        for r in results:
            for element in self._safe_get(r, "missing_elements", []):
                if element in missing_elements_counts:
                    missing_elements_counts[element] += 1

        # Vague terms frequency
        vague_terms_freq = {}
        for r in results:
            for term in self._safe_get(r, "vague_terms_found", []):
                vague_terms_freq[term] = vague_terms_freq.get(term, 0) + 1

        # Multi-control statistics
        multi_control_count = sum(
            1 for r in results if self._safe_get(r, ["multi_control_indicators", "detected"], False))

        # Return all stats as a dictionary
        return {
            "total_controls": total_controls,
            "excellent_count": excellent_count,
            "good_count": good_count,
            "needs_improvement_count": needs_improvement_count,
            "avg_score": avg_score,
            "missing_elements_counts": missing_elements_counts,
            "vague_terms_freq": vague_terms_freq,
            "multi_control_count": multi_control_count
        }

    def _create_report_workbook(self, report_data: Dict, summary_stats: Dict,
                                include_frequency: bool, include_control_type: bool,
                                include_risk_alignment: bool) -> Workbook:
        """
        Create Excel workbook for the report.

        Args:
            report_data: Dictionary containing DataFrames
            summary_stats: Dictionary of summary statistics
            include_frequency: Whether to include frequency validation
            include_control_type: Whether to include control type validation
            include_risk_alignment: Whether to include risk alignment analysis

        Returns:
            Excel workbook object
        """
        # Extract DataFrames
        df_results = report_data["df_results"]
        df_keywords = report_data["df_keywords"]
        df_feedback = report_data["df_feedback"]

        # Extract summary stats
        total_controls = summary_stats["total_controls"]
        excellent_count = summary_stats["excellent_count"]
        good_count = summary_stats["good_count"]
        needs_improvement_count = summary_stats["needs_improvement_count"]
        avg_score = summary_stats["avg_score"]
        missing_elements_counts = summary_stats["missing_elements_counts"]
        vague_terms_freq = summary_stats["vague_terms_freq"]
        multi_control_count = summary_stats["multi_control_count"]

        # Calculate validation statistics if needed
        freq_validation_stats = None
        control_type_validation_stats = None

        if include_frequency:
            valid_freq_count = sum(
                1 for r in df_results["Frequency Valid"] if r == "Yes")
            freq_validation_stats = {
                "Valid": valid_freq_count,
                "Invalid": total_controls - valid_freq_count,
                "Percent Valid": (valid_freq_count / total_controls * 100) if total_controls > 0 else 0
            }

        if include_control_type:
            valid_type_count = sum(
                1 for r in df_results["Control Type Valid"] if r == "Yes")
            control_type_validation_stats = {
                "Valid": valid_type_count,
                "Invalid": total_controls - valid_type_count,
                "Percent Valid": (valid_type_count / total_controls * 100) if total_controls > 0 else 0
            }

        # Create workbook
        wb = Workbook()

        # Analysis Results sheet
        ws_results = wb.active
        ws_results.title = "Analysis Results"

        for r_idx, row in enumerate(dataframe_to_rows(df_results, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_results.cell(row=r_idx, column=c_idx, value=value)

        # Keywords sheet
        ws_keywords = wb.create_sheet(title="Keyword Matches")

        for r_idx, row in enumerate(dataframe_to_rows(df_keywords, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_keywords.cell(row=r_idx, column=c_idx, value=value)

        # Feedback sheet (potentially problematic)
        ws_feedback = wb.create_sheet(title="Enhancement Feedback")

        # Use a simplified approach to write feedback data
        # First, write the headers
        for c_idx, header in enumerate(df_feedback.columns, 1):
            ws_feedback.cell(row=1, column=c_idx, value=header)

        # Then write each row carefully
        for r_idx, record in enumerate(df_feedback.to_dict('records'), 2):
            for c_idx, (column, value) in enumerate(record.items(), 1):
                # Extra safety for each cell
                if column == "Control ID":
                    # Don't sanitize control IDs
                    ws_feedback.cell(row=r_idx, column=c_idx, value=value)
                else:
                    # Sanitize all other values
                    safe_value = self._safe_text(value, 1000)
                    ws_feedback.cell(row=r_idx, column=c_idx, value=safe_value)

        # Multi-Control Candidates sheet (if any detected)
        self._add_multi_control_sheet(wb, multi_control_count)

        # Executive Summary sheet
        ws_summary = wb.create_sheet(title="Executive Summary")

        # Prepare summary data
        summary_data = self._prepare_summary_data(
            summary_stats, freq_validation_stats, control_type_validation_stats
        )

        # Write summary data
        for r_idx, row in enumerate(summary_data, 1):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        # Methodology sheet
        ws_method = wb.create_sheet(title="Methodology")

        # Prepare methodology text
        methodology_text = self._get_methodology_text(
            include_frequency, include_control_type, include_risk_alignment
        )

        # Write methodology data
        for r_idx, row in enumerate(methodology_text, 1):
            for c_idx, value in enumerate(row, 1):
                ws_method.cell(row=r_idx, column=c_idx, value=value)

        # Example Controls sheet
        ws_examples = wb.create_sheet(title="Example Controls")

        # Prepare examples text
        examples_text = self._get_examples_text()

        # Write examples data
        for r_idx, row in enumerate(examples_text, 1):
            for c_idx, value in enumerate(row, 1):
                ws_examples.cell(row=r_idx, column=c_idx, value=value)

        return wb

    def _add_multi_control_sheet(self, wb: Workbook, multi_control_count: int):
        """
        Add a Multi-Control Candidates sheet to the workbook if applicable.

        Args:
            wb: Excel workbook
            multi_control_count: Number of multi-controls detected
        """
        if multi_control_count <= 0:
            return

        from openpyxl.styles import Font, PatternFill

        ws_multi = wb.create_sheet(title="Multi-Control Candidates")

        # Headers
        headers = ["Control ID", "Potential Control", "Score", "Confidence", "Action"]
        for c_idx, header in enumerate(headers, 1):
            cell = ws_multi.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        # Set column widths
        ws_multi.column_dimensions['A'].width = 15
        ws_multi.column_dimensions['B'].width = 60
        ws_multi.column_dimensions['C'].width = 10
        ws_multi.column_dimensions['D'].width = 15
        ws_multi.column_dimensions['E'].width = 30

    def _prepare_summary_data(self, summary_stats: Dict,
                              freq_validation_stats: Optional[Dict],
                              control_type_validation_stats: Optional[Dict]) -> List[List]:
        """
        Prepare summary data for the Executive Summary sheet.

        Args:
            summary_stats: Dictionary of summary statistics
            freq_validation_stats: Frequency validation statistics
            control_type_validation_stats: Control type validation statistics

        Returns:
            List of row data for the summary sheet
        """
        # Extract statistics
        total_controls = summary_stats["total_controls"]
        excellent_count = summary_stats["excellent_count"]
        good_count = summary_stats["good_count"]
        needs_improvement_count = summary_stats["needs_improvement_count"]
        avg_score = summary_stats["avg_score"]
        missing_elements_counts = summary_stats["missing_elements_counts"]
        vague_terms_freq = summary_stats["vague_terms_freq"]
        multi_control_count = summary_stats["multi_control_count"]

        # Summary data
        summary_data = [
            ["Control Description Analysis - Executive Summary", ""],
            ["", ""],
            ["Total Controls Analyzed", total_controls],
            ["Average Score", f"{avg_score:.1f}"],
            ["", ""],
            ["Category Breakdown", ""],
            [f"Excellent ({self.excellent_threshold}-100)", excellent_count],
            [f"Good ({self.good_threshold}-{self.excellent_threshold - 1})", good_count],
            [f"Needs Improvement (0-{self.good_threshold - 1})", needs_improvement_count],
            ["", ""],
            ["Multi-Control Descriptions", multi_control_count],
            ["", ""],
            ["Missing Elements", ""],
        ]

        # Add missing elements stats
        for element, count in missing_elements_counts.items():
            percentage = (count / total_controls) * 100 if total_controls > 0 else 0
            summary_data.append([f"Missing {element}", f"{count} ({percentage:.1f}%)"])

        summary_data.append(["", ""])
        summary_data.append(["Top Vague Terms", ""])

        # Add top vague terms
        sorted_vague_terms = sorted(vague_terms_freq.items(), key=lambda x: x[1], reverse=True)
        for term, count in sorted_vague_terms[:10]:  # Top 10 vague terms
            percentage = (count / total_controls) * 100 if total_controls > 0 else 0
            summary_data.append([term, f"{count} ({percentage:.1f}%)"])

        # Add validation stats
        if freq_validation_stats:
            summary_data.append(["", ""])
            summary_data.append(["Frequency Validation", ""])
            summary_data.append(["Valid Frequency",
                                 f"{freq_validation_stats['Valid']} ({freq_validation_stats['Percent Valid']:.1f}%)"])
            summary_data.append(["Invalid Frequency",
                                 f"{freq_validation_stats['Invalid']} ({100 - freq_validation_stats['Percent Valid']:.1f}%)"])

        if control_type_validation_stats:
            summary_data.append(["", ""])
            summary_data.append(["Control Type Validation", ""])
            summary_data.append(["Valid Control Type",
                                 f"{control_type_validation_stats['Valid']} ({control_type_validation_stats['Percent Valid']:.1f}%)"])
            summary_data.append(["Invalid Control Type",
                                 f"{control_type_validation_stats['Invalid']} ({100 - control_type_validation_stats['Percent Valid']:.1f}%)"])

        return summary_data

    def _format_workbook_sheets(self, wb: Workbook):
        """
        Apply consistent formatting to all sheets in the workbook.

        Args:
            wb: Excel workbook to format
        """
        from openpyxl.styles import Font, PatternFill, Alignment

        # Get sheet references
        ws_results = wb["Analysis Results"]
        ws_keywords = wb["Keyword Matches"]
        ws_feedback = wb["Enhancement Feedback"]
        ws_summary = wb["Executive Summary"]
        ws_method = wb["Methodology"]
        ws_examples = wb["Example Controls"]

        # Format headers for data sheets
        for ws in [ws_results, ws_keywords, ws_feedback]:
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_font = Font(bold=True)

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

        # Set reasonable column widths for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value:
                        try:
                            max_length = max(max_length, min(len(str(cell.value)), 50))
                        except (TypeError, ValueError):
                            pass

                # Cap width at 50
                adjusted_width = min((max_length + 2) * 1.1, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Additional formatting for summary sheet
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20

        # Set fonts for various sections
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True)

        # Format summary sheet titles
        ws_summary.cell(row=1, column=1).font = title_font

        # Format methodology sheet title
        ws_method.column_dimensions['A'].width = 60
        ws_method.cell(row=1, column=1).font = title_font

        # Format examples sheet title
        ws_examples.column_dimensions['A'].width = 70
        ws_examples.cell(row=1, column=1).font = title_font

        # Format section headers in summary
        section_rows = [6, 11, 13]  # Key section start rows
        for row in section_rows:
            if row <= ws_summary.max_row:
                ws_summary.cell(row=row, column=1).font = section_font

        # Format section headers in methodology
        method_section_rows = [3, 14, 24, 39]  # Key section start rows
        for row in method_section_rows:
            if row <= ws_method.max_row:
                ws_method.cell(row=row, column=1).font = section_font

        # Format section headers in examples
        for row in [3, 6, 9, 12]:  # Key section start rows
            ws_examples.cell(row=row, column=1).font = section_font

    def _save_workbook(self, wb: Workbook, output_file: str) -> bool:
        """
        Save workbook with error handling and diagnostics.

        Args:
            wb: Excel workbook to save
            output_file: Path to save the workbook

        Returns:
            Boolean indicating success
        """
        import traceback

        try:
            print(f"🔄 Attempting to save Excel file to: {output_file}")
            logger.info(f"Attempting to save workbook to: {output_file}")

            # Check for dangerous content length in cells
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and len(cell.value) > 32000:
                            logger.warning(
                                f"⚠️ Truncating long string in sheet '{ws.title}' at row {cell.row}, col {cell.column}"
                            )
                            cell.value = cell.value[:32000] + "..."

            wb.save(output_file)
            logger.info(f"✅ Successfully saved workbook to: {output_file}")
            return True

        except Exception as e:
            logger.error(f"❌ Failed to save workbook to {output_file}: {e}")
            traceback.print_exc()

            # Try saving with a backup filename
            try:
                backup_file = os.path.splitext(output_file)[0] + "_backup.xlsx"
                wb.save(backup_file)
                logger.info(f"📁 Saved backup workbook to: {backup_file}")
                return True
            except Exception as backup_error:
                logger.error(f"❌ Failed to save backup workbook: {backup_error}")
                traceback.print_exc()
                return False

    # Helper methods for the report generator
    def _safe_get(self, obj: Dict, key: Union[str, List[str]], default: Any = None) -> Any:
        """
        Safely get a value from an object, handling nested keys and various error conditions.

        Args:
            obj: Dictionary to get value from
            key: String key or list of nested keys
            default: Default value if key not found

        Returns:
            Value from dictionary or default
        """
        if obj is None:
            return default

        # Handle list of keys for nested access
        if isinstance(key, list):
            current = obj
            for k in key:
                if isinstance(current, dict) and k in current:
                    current = current[k]
                else:
                    return default
            return current

        # Simple key access
        if isinstance(obj, dict) and key in obj:
            return obj[key]

        return default

    def _safe_text(self, text: Any, max_length: int = 255) -> str:
        """
        Make text safe for Excel by handling various issues.

        Args:
            text: Text to make safe
            max_length: Maximum length to allow

        Returns:
            Safe text string
        """
        if text is None:
            return ""

        # Convert to string
        if not isinstance(text, str):
            text = str(text)

        # Truncate if too long
        if len(text) > max_length:
            text = text[:max_length - 3] + "..."

        # Prevent formula injection
        if text and text[0] in ['=', '+', '-', '@']:
            text = "'" + text

        # Remove problematic characters
        bad_chars = ['\0', '\x1a', '\r', '\x03']  # NULL, EOF, CR, etc.
        for char in bad_chars:
            text = text.replace(char, '')

        # Replace newlines with spaces to prevent XML issues
        text = text.replace('\n', ' ')

        return text

    def _join_keywords(self, keywords: List[str], separator: str = ", ") -> str:
        """
        Safely join keywords into a string.

        Args:
            keywords: List of keywords to join
            separator: Separator to use

        Returns:
            Joined keywords string
        """
        if not keywords:
            return "None"
        return separator.join(str(kw) for kw in keywords)

    def _get_methodology_text(self, include_frequency: bool,
                              include_control_type: bool,
                              include_risk_alignment: bool) -> List[List[str]]:
        """
        Generate the methodology text for the report.

        Args:
            include_frequency: Whether to include frequency validation
            include_control_type: Whether to include control type validation
            include_risk_alignment: Whether to include risk alignment analysis

        Returns:
            List of row data for the methodology sheet
        """
        methodology_text = [
            ["Enhanced Control Description Analysis Methodology", ""],
            ["", ""],
            ["Overview", ""],
            [
                "This analysis evaluates control descriptions based on five key elements that should be "
                "present in a well-written control description:",
                ""],
            ["", ""],
            ["1. WHO performs the control", ""],
            ["2. WHEN the control is performed", ""],
            ["3. WHAT activities are performed", ""],
            ["4. WHY the control exists", ""],
            ["5. ESCALATION procedures", ""],
            ["", ""],
            ["Enhanced NLP Analysis", ""],
            ["The analysis uses advanced Natural Language Processing (NLP) techniques to:", ""],
            ["", ""],
            ["- Detect exact keyword matches and identify semantically similar terms", ""],
            ["- Analyze the context, relevance, and strength of control actions", ""],
            ["- Identify primary versus secondary performers of the control", ""],
            ["- Detect vague timing terms and suggest specific alternatives", ""],
            ["- Evaluate alignment between control purpose and mapped risks", ""],
            ["- Detect when a description contains multiple controls", ""],
            ["", ""],
            ["Scoring Method", ""],
            ["Each element is weighted based on its importance:", ""],
            ["", ""],
            [f"WHO: {self.elements['WHO'].weight}%", ""],
            [f"WHEN: {self.elements['WHEN'].weight}%", ""],
            [f"WHAT: {self.elements['WHAT'].weight}%", ""],
            [f"WHY: {self.elements['WHY'].weight}%", ""],
            [f"ESCALATION: {self.elements['ESCALATION'].weight}%", ""],
            ["", ""],
            ["Penalties are applied for:", ""],
            [f"- Vague terms (up to {self.max_vague_penalty} points)", ""],
            [f"- Multiple controls in one description (up to {self.max_multi_control_penalty} points)", ""],
            ["", ""],
            ["Categories", ""],
            ["Controls are categorized based on their total score:", ""],
            ["", ""],
            [f"Excellent: {self.excellent_threshold}-100", ""],
            [f"Good: {self.good_threshold}-{self.excellent_threshold - 1}", ""],
            [f"Needs Improvement: 0-{self.good_threshold - 1}", ""]
        ]

        # Add multi-control detection methodology
        methodology_text.append(["", ""])
        methodology_text.append(["Enhanced Multi-Control Detection", ""])
        methodology_text.append(["", ""])
        methodology_text.append([
            "The system uses relationship analysis between elements to identify "
            "when a description contains multiple controls:",
            ""])
        methodology_text.append(["", ""])
        methodology_text.append([
            "- Detects different timing patterns for actions (e.g., daily vs. monthly activities)", ""])
        methodology_text.append(["- Identifies different performers for distinct actions", ""])
        methodology_text.append(["- Recognizes sequence markers that indicate separate procedures", ""])
        methodology_text.append(["- Differentiates between actual multiple controls and escalation paths", ""])
        methodology_text.append(["", ""])
        methodology_text.append([
            "The detection provides confidence levels (high, medium, low) to indicate certainty", ""])

        # Add validation methodology if applicable
        if include_frequency or include_control_type or include_risk_alignment:
            methodology_text.append(["", ""])
            methodology_text.append(["Enhanced Validation Checks", ""])

        if include_frequency:
            methodology_text.append(["", ""])
            methodology_text.append(["Frequency Validation:", ""])
            methodology_text.append([
                "Advanced detection of timing patterns with specific and vague term identification.", ""])
            methodology_text.append(["Validation against frequency metadata with context-aware scoring.", ""])

        if include_control_type:
            methodology_text.append(["", ""])
            methodology_text.append(["Control Type Validation:", ""])
            methodology_text.append([
                "Validates consistency between control actions and declared control type.", ""])
            methodology_text.append([
                "Provides context-specific feedback for preventive, detective, "
                "corrective, and automated controls.",
                ""])

        if include_risk_alignment:
            methodology_text.append(["", ""])
            methodology_text.append(["Risk Alignment:", ""])
            methodology_text.append([
                "Evaluates how well the control's purpose aligns with the mapped risk.", ""])
            methodology_text.append([
                "Identifies both explicit and implicit WHY statements and their strength.", ""])

        return methodology_text

    def _get_examples_text(self) -> List[List[str]]:
        """
        Generate the examples text for the report.

        Returns:
            List of row data for the examples sheet
        """
        return [
            ["Example Controls", ""],
            ["", ""],
            ["Excellent Control Example", ""],
            [
                "The Accounting Manager reviews the monthly reconciliation between the subledger and "
                "general ledger by the 5th business day of the following month. The reviewer examines "
                "supporting documentation, verifies that all reconciling items have been properly "
                "identified and resolved, and ensures compliance with accounting policies. The review "
                "is evidenced by electronic sign-off in the financial system. Any discrepancies "
                "exceeding $10,000 are escalated to the Controller and documented in the issue tracking "
                "system. The reconciliation and review documentation are stored in the Finance SharePoint "
                "site and retained according to the document retention policy.",
                ""],
            ["", ""],
            ["Good Control Example", ""],
            [
                "The Accounting Supervisor reviews the monthly journal entries prior to posting to "
                "ensure accuracy and completeness. The reviewer checks supporting documentation and "
                "approves entries by signing the journal entry form. Any errors are returned to the "
                "preparer for correction.",
                ""],
            ["", ""],
            ["Needs Improvement Control Example", ""],
            ["Management reviews financial statements periodically and addresses any issues as appropriate.", ""],
            ["", ""],
            ["Key Improvement Recommendations", ""],
            ["1. Specify WHO is responsible (exact role, not just 'management')", ""],
            ["2. Define WHEN the control occurs (specific timing, not just 'periodically')", ""],
            ["3. Detail WHAT actions are taken (specific verifications, not just 'reviews')", ""],
            ["4. Explain WHY the control exists (what risk it addresses)", ""],
            ["5. Include ESCALATION procedures for exceptions", ""],
            ["6. Avoid vague terms like 'appropriate', 'timely', 'periodically'", ""],
            ["7. Separate multiple controls into individual control descriptions", ""]
        ]


def setup_logging(log_file: Optional[str] = None, verbose: bool = False) -> None:
    """
    Set up logging configuration for the application.

    Args:
        log_file: Optional path to log file
        verbose: Whether to use verbose (DEBUG) logging
    """
    # Determine log level
    log_level = logging.DEBUG if verbose else logging.INFO

    # Configure root logger
    root_logger = logging.getLogger()
    root_logger.setLevel(log_level)

    # Create formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # Add console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    root_logger.addHandler(console_handler)

    # Add file handler if log file specified
    if log_file:
        try:
            file_handler = logging.FileHandler(log_file)
            file_handler.setFormatter(formatter)
            root_logger.addHandler(file_handler)
            logging.info(f"Logging to file: {log_file}")
        except Exception as e:
            logging.warning(f"Could not set up file logging: {e}")

    # Set module logger level
    logger.setLevel(log_level)


def main() -> int:
    """
    Command-line interface for the Enhanced Control Description Analyzer.

    Returns:
        Exit code (0 for success, 1 for error)
    """
    parser = argparse.ArgumentParser(
        description='Analyze control descriptions with enhanced NLP capabilities.')

    # Input/output arguments
    parser.add_argument('file', nargs='?', help='Excel file with control descriptions')
    parser.add_argument('--output-file', help='Output Excel file path')
    parser.add_argument('--config', help='Path to configuration file')

    # Column mapping arguments
    parser.add_argument('--id-column', help='Column containing control IDs (name)')
    parser.add_argument('--desc-column', help='Column containing control descriptions (name)')
    parser.add_argument('--freq-column', help='Column containing frequency values for validation (name)')
    parser.add_argument('--type-column', help='Column containing control type values for validation (name)')
    parser.add_argument('--risk-column', help='Column containing risk descriptions for WHY alignment (name)')

    # Processing mode arguments
    parser.add_argument('--disable-enhanced-detection', action='store_true',
                        help='Disable enhanced detection modules')
    parser.add_argument('--batch-size', type=int, default=500,
                        help='Number of controls to process in each batch')
    parser.add_argument('--use-batches', action='store_true',
                        help='Use batch processing for large files')
    parser.add_argument('--temp-dir', help='Directory to store temporary batch results')

    # Logging arguments
    parser.add_argument('--log-file', help='Path to log file')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose logging')

    args = parser.parse_args()

    # Set up logging first
    setup_logging(args.log_file, args.verbose)

    # Log startup information
    logger.info("Control Analyzer starting")
    logger.info(f"Python version: {sys.version}")
    if args.file:
        logger.info(f"Input file: {args.file}")

    # Set default output filename if not provided
    if args.file and not args.output_file:
        base_name = os.path.splitext(args.file)[0]
        args.output_file = f"{base_name}_enhanced_analysis.xlsx"
        logger.debug(f"Setting default output file to {args.output_file}")

    # Make sure we have a file to process
    if not args.file:
        logger.error("No input file specified")
        parser.print_help()
        return 1

    try:
        # Create enhanced analyzer with config file if specified
        analyzer = EnhancedControlAnalyzer(args.config)

        # Toggle enhanced detection if requested
        if args.disable_enhanced_detection:
            analyzer.use_enhanced_detection = False
            logger.info("Enhanced detection modules disabled. Using base analysis only.")

        # Process the file
        if args.use_batches:
            # Use batch processing for large files
            analyzer.analyze_file_with_batches(
                args.file,
                args.id_column,
                args.desc_column,
                args.freq_column,
                args.type_column,
                args.risk_column,
                args.output_file,
                args.batch_size,
                args.temp_dir
            )
        else:
            # Use standard processing
            analyzer.analyze_file(
                args.file,
                args.id_column,
                args.desc_column,
                args.freq_column,
                args.type_column,
                args.risk_column,
                args.output_file
            )

        logger.info("Analysis completed successfully")
        return 0

    except Exception as e:
        logger.error(f"Error during analysis: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())