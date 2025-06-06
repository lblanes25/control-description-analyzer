#!/usr/bin/env python3
"""
Validate Spans Script

This script validates that the spans in the review template match the text in the description.
It's useful for checking the accuracy of spans before converting to spaCy format.

Usage:
    python validate_spans.py reviewed_template.xlsx [--output validation_report.xlsx]

Input:
    - Excel file with columns: Control ID, Description, WHO Spans, WHAT Spans, etc.

Output:
    - Validation report (printed or saved to Excel)
"""

import os
import sys
import pandas as pd
import numpy as np
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def excel_format_to_spans(formatted_str):
    """
    Convert Excel-formatted spans back to list of tuples.

    Args:
        formatted_str (str): Formatted string from Excel

    Returns:
        list: List of (keyword, start, end) tuples
    """
    if pd.isna(formatted_str) or formatted_str.strip() == "":
        return []

    spans = []
    for span_str in formatted_str.split(";"):
        span_str = span_str.strip()
        if not span_str:
            continue

        parts = span_str.split("|")
        if len(parts) >= 3:
            keyword = parts[0]
            try:
                start = int(parts[1])
                end = int(parts[2])
                spans.append((keyword, start, end))
            except ValueError:
                # Skip if start/end aren't valid integers
                spans.append((keyword, -1, -1))  # Mark as invalid

    return spans


def validate_span(text, keyword, start, end):
    """
    Validate that the span matches the keyword.

    Args:
        text (str): The full text
        keyword (str): The expected keyword
        start (int): Start index
        end (int): End index

    Returns:
        tuple: (is_valid, actual_text, reason)
    """
    if pd.isna(text):
        return False, "", "Description is empty"

    if start < 0 or end < 0:
        return False, "", "Invalid span indices"

    if end > len(text):
        return False, "Out of bounds", "Span extends beyond text length"

    actual_text = text[start:end]

    if actual_text == keyword:
        return True, actual_text, "Valid"
    else:
        return False, actual_text, f"Text mismatch: expected '{keyword}'"


def validate_spans(input_file, output_file=None):
    """
    Validate spans in a review template.

    Args:
        input_file (str): Path to reviewed Excel template
        output_file (str, optional): Path to output validation report
    """
    # Read the input Excel file
    df = pd.read_excel(input_file)

    # Check for required columns
    required_cols = ["Control ID", "Description"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Required column '{col}' not found in input file")

    # Elements to process
    elements = ["WHO", "WHAT", "WHEN", "WHY", "ESCALATION"]

    # Find available elements
    available_elements = []
    for element in elements:
        spans_col = f"{element} Spans"
        if spans_col in df.columns:
            available_elements.append(element)

    if not available_elements:
        raise ValueError("No element spans columns found in input file")

    print(f"Found spans columns for elements: {available_elements}")

    # Create validation results
    validation_results = []

    # Process each row
    for i, row in df.iterrows():
        control_id = row["Control ID"]
        description = row["Description"]

        for element in available_elements:
            spans_col = f"{element} Spans"
            spans = excel_format_to_spans(row[spans_col])

            for keyword, start, end in spans:
                is_valid, actual_text, reason = validate_span(description, keyword, start, end)

                validation_results.append({
                    "Control ID": control_id,
                    "Element": element,
                    "Keyword": keyword,
                    "Start": start,
                    "End": end,
                    "Actual Text": actual_text,
                    "Valid": is_valid,
                    "Reason": reason
                })

    # Create validation DataFrame
    validation_df = pd.DataFrame(validation_results)

    # Print summary
    total_spans = len(validation_df)
    valid_spans = validation_df["Valid"].sum()
    invalid_spans = total_spans - valid_spans

    print(f"\nValidation Results:")
    print(f"Total spans: {total_spans}")
    print(f"Valid spans: {valid_spans} ({valid_spans / total_spans * 100:.1f}%)")
    print(f"Invalid spans: {invalid_spans} ({invalid_spans / total_spans * 100:.1f}%)")

    # Group by element and print
    element_stats = validation_df.groupby("Element")["Valid"].agg(["count", "sum"])
    element_stats["invalid"] = element_stats["count"] - element_stats["sum"]
    element_stats["valid_pct"] = element_stats["sum"] / element_stats["count"] * 100

    print("\nBreakdown by element:")
    for element, stats in element_stats.iterrows():
        print(f"{element}: {stats['sum']}/{stats['count']} valid ({stats['valid_pct']:.1f}%)")

    # Output to Excel if requested
    if output_file:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Results"

        # Define styles
        header_fill = PatternFill(start_color="00CFCFCF", end_color="00CFCFCF", fill_type="solid")
        header_font = Font(bold=True)
        valid_fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
        invalid_fill = PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid")

        # Add header row
        header_row = list(validation_df.columns)
        for col_idx, header in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Add data rows
        for row_idx, result in enumerate(validation_df.itertuples(), 2):
            ws.cell(row=row_idx, column=1, value=result.Control_ID)
            ws.cell(row=row_idx, column=2, value=result.Element)
            ws.cell(row=row_idx, column=3, value=result.Keyword)
            ws.cell(row=row_idx, column=4, value=result.Start)
            ws.cell(row=row_idx, column=5, value=result.End)
            ws.cell(row=row_idx, column=6, value=result.Actual_Text)
            ws.cell(row=row_idx, column=7, value=result.Valid)
            ws.cell(row=row_idx, column=8, value=result.Reason)

            # Apply conditional formatting
            for col_idx in range(1, 9):
                if result.Valid:
                    ws.cell(row=row_idx, column=col_idx).fill = valid_fill
                else:
                    ws.cell(row=row_idx, column=col_idx).fill = invalid_fill

        # Auto-adjust column widths
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    text_length = len(str(cell.value))
                    if text_length > max_length:
                        max_length = text_length

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")

        # Add summary data
        summary_data = [
            ["Validation Summary", ""],
            ["Total spans", total_spans],
            ["Valid spans", valid_spans],
            ["Invalid spans", invalid_spans],
            ["Valid percentage", f"{valid_spans / total_spans * 100:.1f}%"],
            ["", ""],
            ["Element", "Total", "Valid", "Invalid", "Valid %"]
        ]

        for element, stats in element_stats.iterrows():
            summary_data.append([
                element,
                stats["count"],
                stats["sum"],
                stats["invalid"],
                f"{stats['valid_pct']:.1f}%"
            ])

        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == 7:
                    cell.font = header_font
                    if col_idx == 1:
                        cell.fill = header_fill

        # Save workbook
        wb.save(output_file)
        print(f"\nValidation report saved to {output_file}")

    # Return validation DataFrame
    return validation_df


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description="Validate spans in a reviewed Excel template")
    parser.add_argument("input_file", help="Input Excel template with reviewed spans")
    parser.add_argument("--output", help="Output validation report Excel file")
    return parser.parse_args()


def main():
    """Main function"""
    args = parse_arguments()
    validate_spans(args.input_file, args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())