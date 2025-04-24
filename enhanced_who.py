import re
import spacy
from spacy.matcher import PhraseMatcher
from spacy.tokens import Span
from typing import Dict, List, Any, Optional, Tuple, Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def identify_control_action_subjects(doc):
    """Identify the true subjects performing control actions"""
    control_verbs = [
        "review", "approve", "verify", "check", "validate", "reconcile",
        "examine", "analyze", "evaluate", "assess", "monitor", "track",
        "investigate", "inspect", "audit", "oversee", "supervise", "ensure",
        "perform", "execute", "conduct", "disable", "enforce", "generate"
    ]

    subjects = []

    for token in doc:
        # Check if token is a control verb
        if token.lemma_ in control_verbs:
            # Find the subject of this verb
            for child in token.children:
                if child.dep_ in ("nsubj", "nsubjpass"):
                    # Get the full noun phrase for this subject
                    for chunk in doc.noun_chunks:
                        if child.i >= chunk.start and child.i < chunk.end:
                            subjects.append({
                                "text": chunk.text,
                                "verb": token.text,
                                "verb_lemma": token.lemma_,
                                "is_passive": child.dep_ == "nsubjpass",
                                "start": chunk.start,
                                "end": chunk.end
                            })
                            break

    return subjects


def classify_entity_type(text, nlp):
    """Classify an entity as human, system, or non-performer"""
    text_lower = text.lower()

    human_indicators = [
        "manager", "director", "supervisor", "analyst", "specialist", "officer",
        "coordinator", "lead", "team", "staff", "personnel", "employee",
        "individual", "person", "accountant", "controller", "auditor", "administrator"
    ]

    system_indicators = [
        "system", "application", "software", "platform", "database", "server",
        "program", "script", "job", "batch", "workflow", "algorithm", "automated"
    ]

    non_performer_indicators = [
        "limit", "threshold", "policy", "procedure", "standard", "regulation",
        "account", "transaction", "balance", "report", "document", "record"
    ]

    # Check if the entity contains human role indicators
    if any(indicator in text_lower for indicator in human_indicators):
        return "human"

    # Check if the entity contains system indicators
    if any(indicator in text_lower for indicator in system_indicators):
        return "system"

    # Check if the entity contains non-performer indicators
    if any(indicator in text_lower for indicator in non_performer_indicators):
        return "non-performer"

    # Use NLP to check for person or organization entities
    doc = nlp(text)
    for ent in doc.ents:
        if ent.label_ in ("PERSON", "ORG"):
            return "human"

    # Default to unknown if no clear classification
    return "unknown"


def calculate_who_confidence(entity, control_type=None, frequency=None):
    """Calculate confidence score for a WHO entity"""
    base_score = 0.5  # Start with neutral score

    # Adjust for entity type
    if entity["type"] == "human":
        base_score += 0.2
    elif entity["type"] == "system":
        base_score += 0.1
    elif entity["type"] == "non-performer":
        base_score = 0.1  # Very low confidence for non-performers

    # Adjust for passive vs active voice
    if entity.get("is_passive", False):
        base_score -= 0.2

    # Adjust for control type consistency
    if control_type:
        control_type_lower = control_type.lower()

        if "automated" in control_type_lower and entity["type"] == "system":
            base_score += 0.2
        elif "automated" in control_type_lower and entity["type"] == "human":
            base_score -= 0.2
        elif "manual" in control_type_lower and entity["type"] == "human":
            base_score += 0.2
        elif "manual" in control_type_lower and entity["type"] == "system":
            base_score -= 0.2

    # Adjust for frequency consistency
    if frequency:
        frequency_lower = frequency.lower()

        if any(term in frequency_lower for term in ["daily", "continuous"]):
            if entity["type"] == "system":
                base_score += 0.1
            elif "director" in entity["text"].lower() or "executive" in entity["text"].lower():
                base_score -= 0.1

    # Ensure score is within valid range
    return max(0.1, min(1.0, base_score))


def enhanced_who_detection_v2(text, nlp, control_type=None, frequency=None, existing_keywords=None):
    """
    Enhanced WHO detection with improved subject-verb analysis and entity classification
    """
    if not text or text.strip() == '':
        return {
            "primary": None,
            "secondary": [],
            "confidence": 0.2,
            "message": "No text provided"
        }

    try:
        # Process the text
        doc = nlp(text)

        # Find subjects of control actions
        action_subjects = identify_control_action_subjects(doc)

        # If no action subjects found, check for passive constructions
        if not action_subjects:
            passive_verbs = [token for token in doc if token.dep_ == "auxpass"]
            if passive_verbs:
                return {
                    "primary": {
                        "text": "Unknown Performer",
                        "verb": passive_verbs[0].head.text if passive_verbs else "unknown",
                        "type": "unknown",
                        "score": 0.2
                    },
                    "secondary": [],
                    "confidence": 0.2,
                    "message": "Passive voice detected - no clear performer specified"
                }

        # Classify and score all potential WHO entities
        who_candidates = []

        for subject in action_subjects:
            entity_type = classify_entity_type(subject["text"], nlp)

            # Skip non-performers
            if entity_type == "non-performer":
                continue

            # Calculate confidence score
            confidence = calculate_who_confidence(
                {"text": subject["text"], "type": entity_type, "is_passive": subject.get("is_passive", False)},
                control_type,
                frequency
            )

            who_candidates.append({
                "text": subject["text"],
                "verb": subject["verb"],
                "type": entity_type,
                "score": confidence,
                "position": subject["start"]
            })

        # If no valid candidates found, provide a fallback
        if not who_candidates:
            # Look for any noun phrases that could be performers
            for chunk in doc.noun_chunks:
                entity_type = classify_entity_type(chunk.text, nlp)

                # Skip non-performers
                if entity_type == "non-performer":
                    continue

                # Calculate a lower confidence score for fallback
                confidence = calculate_who_confidence(
                    {"text": chunk.text, "type": entity_type},
                    control_type,
                    frequency
                ) * 0.7  # Apply penalty for not being linked to a control verb

                who_candidates.append({
                    "text": chunk.text,
                    "verb": "unknown",
                    "type": entity_type,
                    "score": confidence,
                    "position": chunk.start
                })

        # If still no candidates, return unknown
        if not who_candidates:
            return {
                "primary": {
                    "text": "Unknown Performer",
                    "verb": "unknown",
                    "type": "unknown",
                    "score": 0.2
                },
                "secondary": [],
                "confidence": 0.2,
                "message": "No performer detected"
            }

        # Sort candidates by score and then by position in text
        who_candidates.sort(key=lambda x: (-x["score"], x["position"]))

        # Generate message
        message = ""

        # Check for consistency with control type
        if control_type and who_candidates[0]["type"] != "unknown":
            control_type_lower = control_type.lower()
            if "automated" in control_type_lower and who_candidates[0]["type"] == "human":
                message = "Warning: Human performer detected for automated control"
            elif "manual" in control_type_lower and who_candidates[0]["type"] == "system":
                message = "Warning: System performer detected for manual control"

        # Check for vague performers
        if who_candidates[0]["text"].lower() in ["management", "staff", "team"]:
            message = "Warning: Vague performer detected - consider specifying exact role"

        # Identify primary and secondary performers
        primary = who_candidates[0]
        secondary = who_candidates[1:] if len(who_candidates) > 1 else []

        return {
            "primary": primary,
            "secondary": secondary,
            "confidence": primary["score"],
            "message": message
        }

    except Exception as e:
        print(f"Error in enhanced WHO detection: {str(e)}")
        return {
            "primary": None,
            "secondary": [],
            "confidence": 0.2,
            "message": f"Error: {str(e)}"
        }


def generate_who_detection_report(results, output_file):
    """Generate a detailed WHO detection report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "WHO Detection Analysis"

    # Define headers
    headers = [
        "Control ID",
        "Description",
        "Primary WHO",
        "Entity Type",
        "Action Verb",
        "Confidence",
        "Secondary WHO",
        "Message",
        "Improvement Recommendation"
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Apply formatting
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Write data
    row_idx = 2
    for r in results:
        who_detection = r.get("who_detection", {})
        primary = who_detection.get("primary", {})
        secondary = who_detection.get("secondary", [])

        # Format description for better readability
        description = r["description"]
        if len(description) > 100:
            description = description[:97] + "..."

        # Generate improvement recommendation
        recommendation = ""
        if primary.get("text") == "Unknown Performer":
            recommendation = "Specify who performs this control action"
        elif primary.get("score", 0) < 0.5:
            recommendation = "Clarify the performer of this control"
        elif who_detection.get("message", "").startswith("Warning"):
            if "vague" in who_detection.get("message", "").lower():
                recommendation = "Replace vague role with specific position title"
            elif "automated" in who_detection.get("message", "").lower():
                recommendation = "Review control type classification or clarify performer"

        # Format secondary performers
        secondary_text = ""
        if secondary:
            secondary_text = "; ".join([f"{s['text']} ({s['type']}, {s['score']:.2f})" for s in secondary[:2]])
            if len(secondary) > 2:
                secondary_text += f"; +{len(secondary) - 2} more"

        # Write row data
        ws.cell(row=row_idx, column=1, value=r["control_id"])
        ws.cell(row=row_idx, column=2, value=description)
        ws.cell(row=row_idx, column=3, value=primary.get("text", "None"))
        ws.cell(row=row_idx, column=4, value=primary.get("type", "unknown"))
        ws.cell(row=row_idx, column=5, value=primary.get("verb", "unknown"))
        ws.cell(row=row_idx, column=6, value=f"{primary.get('score', 0):.2f}")
        ws.cell(row=row_idx, column=7, value=secondary_text)
        ws.cell(row=row_idx, column=8, value=who_detection.get("message", ""))
        ws.cell(row=row_idx, column=9, value=recommendation)

        # Apply conditional formatting
        if primary.get("score", 0) < 0.4:
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif who_detection.get("message", "").startswith("Warning"):
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        row_idx += 1

    # Set column widths
    column_widths = {
        1: 15,  # Control ID
        2: 40,  # Description
        3: 25,  # Primary WHO
        4: 15,  # Entity Type
        5: 15,  # Action Verb
        6: 12,  # Confidence
        7: 40,  # Secondary WHO
        8: 40,  # Message
        9: 40  # Recommendation
    }

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Save workbook
    wb.save(output_file)

    return f"WHO detection report saved to {output_file}"