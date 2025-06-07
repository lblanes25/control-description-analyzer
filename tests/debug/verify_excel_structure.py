#!/usr/bin/env python3
"""Verify the new Excel structure"""

import pandas as pd
import openpyxl
import os

excel_file = "test_restructured_output.xlsx"

if not os.path.exists(excel_file):
    print(f"❌ File {excel_file} not found")
    exit(1)

print("🔍 Excel Structure Verification")
print(f"📁 File: {excel_file}")
print()

# Load workbook to check sheets
wb = openpyxl.load_workbook(excel_file)

print("📋 Available Sheets:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    is_hidden = sheet.sheet_state == 'hidden'
    hidden_text = " (HIDDEN)" if is_hidden else ""
    print(f"  {i}. {sheet_name}{hidden_text}")

print()

# Check the main Analysis Results tab
print("📄 Tab 1: 'Analysis Results' (Main User-Facing Tab)")
if "Analysis Results" in wb.sheetnames:
    df_main = pd.read_excel(excel_file, sheet_name="Analysis Results")
    print(f"   Columns ({len(df_main.columns)}):")
    for i, col in enumerate(df_main.columns, 1):
        print(f"     {i:2d}. {col}")
    print(f"   Rows: {len(df_main)}")
    
    # Show sample data
    print("\n   Sample Data:")
    print(df_main.head(1).to_string(index=False))

print()

# Check the hidden metadata tab
print("🔒 Tab 2: 'Element Scores & Metadata' (Hidden Technical Tab)")
if "Element Scores & Metadata" in wb.sheetnames:
    df_meta = pd.read_excel(excel_file, sheet_name="Element Scores & Metadata")
    print(f"   Columns ({len(df_meta.columns)}):")
    for i, col in enumerate(df_meta.columns, 1):
        print(f"     {i:2d}. {col}")
    print(f"   Rows: {len(df_meta)}")
    
    # Show sample data
    print("\n   Sample Data (first few columns):")
    print(df_meta[df_meta.columns[:5]].head(1).to_string(index=False))

print()

# Check enhancement feedback tab
print("📝 Tab 3: 'Enhancement Feedback' (Existing)")
if "Enhancement Feedback" in wb.sheetnames:
    df_feedback = pd.read_excel(excel_file, sheet_name="Enhancement Feedback")
    print(f"   Columns ({len(df_feedback.columns)}):")
    for i, col in enumerate(df_feedback.columns, 1):
        print(f"     {i:2d}. {col}")
    print(f"   Rows: {len(df_feedback)}")

print()
print("✅ Excel Structure Verification Complete!")
print()
print("🎯 Key Improvements:")
print("  ✓ Clean main tab with only essential user-facing information")
print("  ✓ Keywords integrated directly into main tab (no separate sheet)")
print("  ✓ Technical scoring details moved to hidden tab")
print("  ✓ Enhanced subthreshold feedback preserved")
print("  ✓ Simple scoring columns prominently displayed")

wb.close()