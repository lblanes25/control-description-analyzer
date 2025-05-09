#!/usr/bin/env python3
"""
Generate Review Template Script

This script takes the analyzer output Excel file and creates a review template
with start and end positions for each matched keyword entity.

Usage:
    python generate_review_template.py input.xlsx output_review_template.xlsx

Input:
    - Excel file with columns: Control ID, Description, WHO Keywords, WHAT Keywords, etc.

Output:
    - Excel file with additional columns: WHO Start, WHO End, WHAT Start, WHAT End, etc.
    - Each row may have multiple entities for each element type
"""

import os
import sys
import pandas as pd
import numpy as np
import argparse
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.cell import get_column_letter


def find_keyword_spans(text, keywords):
    """
    Find start and end character indices for each keyword in the text.

    Args:
        text (str): The full text to search in
        keywords (str): Comma-separated list of keywords

    Returns:
        list: List of tuples (keyword, start, end)
    """
    if pd.isna(text) or pd.isna(keywords) or keywords.strip() == "None":
        return []

    spans = []
    keywords_list = [kw.strip() for kw in keywords.split(",")]

    for keyword in keywords_list:
        if not keyword:
            continue

        # Find all occurrences of the keyword
        for match in re.finditer(re.escape(keyword), text):
            spans.append((keyword, match.start(), match.end()))

    return spans


def span_to_excel_format(spans):
    """
    Convert spans to a format suitable for Excel:
    - Multiple spans are separated by semicolons
    - Each span is keyword|start|end

    Args:
        spans (list): List of (keyword, start, end) tuples

    Returns:
        str: Formatted string for Excel
    """
    if not spans:
        return ""

    return "; ".join([f"{keyword}|{start}|{end}" for keyword, start, end in spans])


def excel_format_to_spans(formatted_str):
    """
    Convert Excel-formatted spans back to list of tuples.
    Inverse of span_to_excel_format.

    Args:
        formatted_str (str): Formatted string from Excel

    Returns:
        list: List of (keyword, start, end) tuples
    """
    if pd.isna(formatted_str) or formatted_str.strip() == "":
        return []

    spans = []
    for span_str in formatted_str.split(";"):
        span_str = span_str.strip()
        if not span_str:
            continue

        parts = span_str.split("|")
        if len(parts) >= 3:
            keyword = parts[0]
            try:
                start = int(parts[1])
                end = int(parts[2])
                spans.append((keyword, start, end))
            except ValueError:
                # Skip if start/end aren't valid integers
                continue

    return spans


def create_review_template(input_file, output_file):
    """
    Create an Excel review template from analyzer output.

    Args:
        input_file (str): Path to analyzer output Excel file
        output_file (str): Path to output review template Excel file
    """
    # Read the input Excel file
    df = pd.read_excel(input_file)

    # Check for required columns
    required_cols = ["Control ID", "Description"]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"Required column '{col}' not found in input file")

    # Elements to process
    elements = ["WHO", "WHAT", "WHEN", "WHY", "ESCALATION"]

    # Create a new DataFrame for the review template
    review_df = df.copy()

    # Process each element
    for element in elements:
        keyword_col = f"{element} Keywords"

        if keyword_col not in df.columns:
            print(f"Warning: '{keyword_col}' column not found, skipping")
            continue

        spans_col = f"{element} Spans"
        review_df[spans_col] = ""

        # Find spans for each row
        for i, row in df.iterrows():
            description = row["Description"]
            keywords = row[keyword_col]

            spans = find_keyword_spans(description, keywords)
            review_df.at[i, spans_col] = span_to_excel_format(spans)

    # Save the review template
    review_df.to_excel(output_file, index=False)

    # Format the Excel file for easier review
    wb = load_workbook(output_file)
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True)

    # Apply styles to header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Set column widths
    ws.column_dimensions[get_column_letter(df.columns.get_loc("Description") + 1)].width = 60

    for element in elements:
        spans_col = f"{element} Spans"
        if spans_col in df.columns:
            col_idx = df.columns.get_loc(spans_col) + 1
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

    # Save the formatted workbook
    wb.save(output_file)

    print(f"Review template created: {output_file}")
    print(f"Found and processed columns for elements: {[e for e in elements if f'{e} Keywords' in df.columns]}")


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description="Generate review template for control analyzer output")
    parser.add_argument("input_file", help="Input Excel file from control analyzer")
    parser.add_argument("output_file", help="Output Excel review template")
    return parser.parse_args()


def main():
    """Main function"""
    args = parse_arguments()
    create_review_template(args.input_file, args.output_file)
    return 0


if __name__ == "__main__":
    sys.exit(main())