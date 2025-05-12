#!/usr/bin/env python3
"""
Control Description Analyzer - Auditor Review Template Generator

This script generates an auditor-friendly review template from Control Description Analyzer results.
It creates an Excel file formatted with validation dropdowns, color-coding, and clear instructions
that allows non-technical auditors to review the analysis results without needing to understand
the technical details of the analyzer.

Usage:
    python auditor_review_template.py input_file.xlsx output_file.xlsx

Args:
    input_file.xlsx: Excel file containing Control Analyzer results
    output_file.xlsx: Output file path for the auditor review template

Dependencies:
    - pandas
    - openpyxl
"""

import os
import sys
import pandas as pd
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment


def create_auditor_review_template(input_file, output_file, add_instructions=True):
    """
    Create an auditor-friendly review template from Control Analyzer results.

    Args:
        input_file (str): Path to the Excel file containing analyzer results
        output_file (str): Path for the output review template
        add_instructions (bool): Whether to add an instructions sheet

    Returns:
        str: Path to created template file
    """
    print(f"Reading analyzer results from: {input_file}")

    try:
        # Read the input Excel file
        df = pd.read_excel(input_file)

        # Verify that required columns exist
        required_columns = ["Control ID", "Description", "Total Score", "Category"]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            print(f"Warning: Missing required columns: {', '.join(missing_columns)}")
            print("The template will be created but may be missing some information.")

        # Create a review template with simplified columns
        review_df = df[["Control ID"]].copy()

        # Add essential columns when available
        for col in ["Description", "Total Score", "Category", "Missing Elements", "Multiple Controls"]:
            if col in df.columns:
                review_df[col] = df[col]

        # Elements to process
        elements = ["WHO", "WHAT", "WHEN", "WHY", "ESCALATION"]

        # Add element-specific columns
        for element in elements:
            # Add score column if available
            if f"{element} Score" in df.columns:
                review_df[f"{element} Score"] = df[f"{element} Score"]

            # Add keywords column if available
            if f"{element} Keywords" in df.columns:
                # Clean up "None" values
                review_df[f"{element} Keywords"] = df[f"{element} Keywords"].apply(
                    lambda x: "" if pd.isna(x) or x == "None" else x
                )

            # Add validation columns for auditor input
            review_df[f"{element} Correct?"] = ""  # Yes/No/Partial column
            review_df[f"{element} Comments"] = ""  # For auditor feedback

        # Add overall validation columns
        review_df["Overall Score Correct?"] = ""
        review_df["Category Correct?"] = ""
        review_df["Additional Comments"] = ""

        # Add auditor info columns
        review_df["Reviewed By"] = ""
        review_df["Review Date"] = ""

        # Save to Excel
        review_df.to_excel(output_file, index=False, sheet_name="Controls Review")

        # Format the Excel file for better readability
        format_review_template(output_file, elements, add_instructions)

        print(f"Auditor review template created successfully: {output_file}")
        return output_file

    except Exception as e:
        print(f"Error creating review template: {e}")
        raise


def format_review_template(file_path, elements, add_instructions=True):
    """
    Format the Excel file for better readability and usability.

    Args:
        file_path (str): Path to the Excel file to format
        elements (list): List of elements (WHO, WHAT, etc.)
        add_instructions (bool): Whether to add an instructions sheet
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    score_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    keywords_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    validation_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get column mappings
    column_mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col_idx).value
        if col_name:
            column_mapping[col_name] = col_idx

    # Find the index for specific columns
    desc_col = column_mapping.get("Description")

    # Apply styles to headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Apply alternating row colors and borders to data cells
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        row_fill = alt_fill if row % 2 == 0 else None

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            if row_fill:
                cell.fill = row_fill

    # Apply styles based on column type
    for col in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col).value

        # Apply specific styles based on column content
        if col_name:
            # Apply fill to score columns
            if "Score" in col_name and "Correct" not in col_name:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    # Don't override alternating row colors
                    if row % 2 != 0:
                        cell.fill = score_fill

            # Apply fill to keywords columns
            elif "Keywords" in col_name:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    # Don't override alternating row colors
                    if row % 2 != 0:
                        cell.fill = keywords_fill

            # Apply fill to validation columns
            elif "Correct?" in col_name or col_name == "Reviewed By" or col_name == "Review Date":
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    # Don't override alternating row colors
                    if row % 2 != 0:
                        cell.fill = validation_fill

    # Set column widths
    standard_width = 15

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        col_name = ws.cell(row=1, column=col).value

        if col_name == "Description":
            ws.column_dimensions[col_letter].width = 60
        elif "Comments" in str(col_name):
            ws.column_dimensions[col_letter].width = 25
        elif "Keywords" in str(col_name):
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = standard_width

    # Add data validation for "Correct?" columns
    validation = DataValidation(type="list", formula1='"Yes,No,Partially"')
    validation.prompt = "Select Yes, No, or Partially"
    validation.promptTitle = "Validation Options"

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        col_name = ws.cell(row=1, column=col).value

        if col_name and "Correct?" in col_name:
            validation.add(f"{col_letter}2:{col_letter}{ws.max_row}")

            # Add comments to header cells
            header_cell = ws.cell(row=1, column=col)
            comment = Comment(f"Select 'Yes' if the analyzer correctly identified this element.\n"
                              f"Select 'No' if the analyzer missed this element or is wrong.\n"
                              f"Select 'Partially' if the analyzer is somewhat correct but incomplete.", "System")
            header_cell.comment = comment

    ws.add_data_validation(validation)

    # Add reviewer data validation
    date_validation = DataValidation(type="date")
    date_col = column_mapping.get("Review Date")
    if date_col:
        date_letter = get_column_letter(date_col)
        date_validation.add(f"{date_letter}2:{date_letter}{ws.max_row}")
        ws.add_data_validation(date_validation)

    # Freeze panes to keep headers visible
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Add instructions sheet if requested
    if add_instructions:
        add_instructions_sheet(wb, elements)

    # Save the workbook
    wb.save(file_path)


def add_instructions_sheet(workbook, elements):
    """
    Add an instructions sheet to the workbook to guide auditors.

    Args:
        workbook: The openpyxl workbook
        elements (list): List of elements (WHO, WHAT, etc.)
    """
    # Create a new sheet for instructions and position it first
    if "Instructions" in workbook.sheetnames:
        # Remove existing instructions sheet to recreate it
        del workbook["Instructions"]

    ws = workbook.create_sheet("Instructions", 0)

    # Set column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 80

    # Create title and introduction
    ws.cell(row=1, column=1, value="INSTRUCTIONS").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells("A1:B1")

    row = 3
    ws.cell(row=row, column=1, value="Purpose:").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value="This template is designed to help you review the results from the Control Description Analyzer. Your feedback will be used to improve the analyzer's accuracy.")

    row += 2
    ws.cell(row=row, column=1, value="Instructions:").font = Font(bold=True)
    instructions = [
        "1. Review each control description and the analyzer's findings",
        "2. For each element (WHO, WHAT, WHEN, WHY, ESCALATION), indicate if the analyzer is correct",
        "3. Select 'Yes', 'No', or 'Partially' in the '[Element] Correct?' columns",
        "4. Provide comments in the '[Element] Comments' columns, especially for 'No' or 'Partially' responses",
        "5. Review the overall score and category, and indicate if they're correct",
        "6. Enter your name and the review date"
    ]

    for i, instruction in enumerate(instructions):
        ws.cell(row=row + i, column=2, value=instruction)

    row += len(instructions) + 2
    ws.cell(row=row, column=1, value="Elements:").font = Font(bold=True)

    element_descriptions = {
        "WHO": "The person, role, group, or system that performs the control",
        "WHAT": "The specific actions or activities performed as part of the control",
        "WHEN": "The timing or frequency of the control (daily, monthly, etc.)",
        "WHY": "The purpose or objective of the control (what risk it addresses)",
        "ESCALATION": "How exceptions, issues, or failures are handled and escalated"
    }

    for i, element in enumerate(elements):
        if element in element_descriptions:
            ws.cell(row=row + i, column=1, value=element).font = Font(bold=True)
            ws.cell(row=row + i, column=2, value=element_descriptions[element])

    row += len(elements) + 2
    ws.cell(row=row, column=1, value="Examples:").font = Font(bold=True)

    examples = [
        "• Strong WHO example: 'The Finance Manager reviews...' (clear role)",
        "• Weak WHO example: 'Management reviews...' (vague)",
        "",
        "• Strong WHEN example: 'monthly by the 5th business day' (specific timing)",
        "• Weak WHEN example: 'periodically' or 'as needed' (vague)",
        "",
        "• Strong WHAT example: 'reconciles the subledger to the general ledger' (specific action)",
        "• Weak WHAT example: 'reviews the data' (vague)",
        "",
        "• Strong WHY example: 'to ensure accurate financial reporting' (clear purpose)",
        "• Weak WHY example: 'for compliance purposes' (vague)",
        "",
        "• Strong ESCALATION example: 'Exceptions over $10,000 are reported to the Controller' (specific)",
        "• Weak ESCALATION example: 'Issues are addressed as needed' (vague)"
    ]

    for i, example in enumerate(examples):
        ws.cell(row=row + i, column=2, value=example)

    row += len(examples) + 2
    ws.cell(row=row, column=1, value="Validation:").font = Font(bold=True)

    validation_info = [
        "• 'Yes' - The analyzer correctly identified this element",
        "• 'No' - The analyzer incorrectly identified or missed this element",
        "• 'Partially' - The analyzer identified some but not all aspects of this element"
    ]

    for i, info in enumerate(validation_info):
        ws.cell(row=row + i, column=2, value=info)

    row += len(validation_info) + 2
    ws.cell(row=row, column=1, value="Notes:").font = Font(bold=True)
    ws.cell(row=row, column=2,
            value="When providing feedback, please be as specific as possible about what the analyzer missed or incorrectly identified. This will help us improve the analyzer's accuracy.")

    # Auto-fit all rows
    for i in range(1, row + 2):
        ws.row_dimensions[i].height = 15


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description="Generate an auditor-friendly review template from Control Analyzer results"
    )
    parser.add_argument("input_file", help="Excel file with Control Analyzer results")
    parser.add_argument(
        "--output_file", "-o",
        help="Output file for the review template (default: 'input_name_review_template.xlsx')"
    )
    parser.add_argument(
        "--no-instructions", action="store_true",
        help="Skip adding the instructions sheet"
    )

    args = parser.parse_args()

    # If output file not specified, generate one based on input filename
    if not args.output_file:
        input_path = Path(args.input_file)
        stem = input_path.stem
        args.output_file = str(input_path.with_name(f"{stem}_review_template.xlsx"))

    return args


def main():
    """Main entry point"""
    args = parse_arguments()

    try:
        create_auditor_review_template(
            args.input_file,
            args.output_file,
            not args.no_instructions
        )
        return 0
    except Exception as e:
        print(f"Error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())